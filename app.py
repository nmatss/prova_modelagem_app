import os
import logging
from logging.handlers import RotatingFileHandler
from flask import Flask, render_template, request, redirect, url_for, send_from_directory, send_file, flash, jsonify, make_response, abort
from flask_compress import Compress
from weasyprint import HTML, CSS
from werkzeug.datastructures import MultiDict
from werkzeug.utils import secure_filename
# from xhtml2pdf import pisa  # Comentado temporariamente - instalar: pip install xhtml2pdf
from flask_login import LoginManager, login_required, current_user
from auth import auth_bp, get_user_by_id
from admin import admin_bp
from audit_bp import audit_bp  # ✅ Habilitado - AuditLog existe no banco (models.py)
from db import init_app as init_db
from models import db, Relatorio, Referencia, Prova, Foto, AuditLog, Fornecedor, ChecklistTemplate, ChecklistResposta, ArquivoVersao, PreferenciaUsuario, ImportJob, LinkPublico, Manual
from config import Config
from utils import save_file
import json
from excel_export import export_relatorios_to_excel, export_detalhes_to_excel, export_editavel
from error_handlers import register_error_handlers
from security import init_security, SecurityHeaders
from sqlalchemy import desc

# Configurar logging será feito após criar o app
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Carregar configurações do config.py
app.config.from_object(Config)
Config.init_app(app)

# Configurar logging para produção
if not app.debug:
    # Log para arquivo
    if Config.LOG_FILE:
        file_handler = RotatingFileHandler(
            Config.LOG_FILE,
            maxBytes=10485760,  # 10MB
            backupCount=10
        )
        file_handler.setFormatter(logging.Formatter(
            '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
        ))
        file_handler.setLevel(getattr(logging, Config.LOG_LEVEL.upper()))
        app.logger.addHandler(file_handler)

    # Log para console
    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(getattr(logging, Config.LOG_LEVEL.upper()))
    stream_handler.setFormatter(logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    ))
    app.logger.addHandler(stream_handler)

    app.logger.setLevel(getattr(logging, Config.LOG_LEVEL.upper()))
    app.logger.info('Aplicação de Provas iniciada em modo produção')
else:
    # Desenvolvimento - log simples
    logging.basicConfig(level=logging.INFO)
    app.logger.info('Aplicação de Provas iniciada em modo desenvolvimento')

# Inicializar database
init_db(app)

# Inicializar segurança
init_security(app)

# Inicializar i18n (Flask-Babel)
try:
    from flask_babel import Babel, gettext, lazy_gettext

    def _get_locale():
        """Seleciona o locale baseado em (1) ?lang= na URL, (2) preferência salva no usuário, (3) Accept-Language."""
        from flask import request as _req, session as _sess
        from flask_login import current_user as _cu

        lang_param = _req.args.get('lang')
        if lang_param in app.config.get('LANGUAGES', ['pt']):
            _sess['idioma'] = lang_param
            return lang_param

        if 'idioma' in _sess and _sess['idioma'] in app.config.get('LANGUAGES', ['pt']):
            return _sess['idioma']

        if _cu and _cu.is_authenticated and getattr(_cu, 'idioma', None) in app.config.get('LANGUAGES', ['pt']):
            return _cu.idioma

        return _req.accept_languages.best_match(app.config.get('LANGUAGES', ['pt'])) or 'pt'

    babel = Babel(app, locale_selector=_get_locale)

    # Expor gettext/_l em templates Jinja
    app.jinja_env.globals['_'] = gettext
    app.jinja_env.globals['_l'] = lazy_gettext
except ImportError:
    # Flask-Babel ainda não instalado — sistema continua em PT fixo
    app.logger.warning('Flask-Babel não instalado. Site continuará em PT até instalar.')
    app.jinja_env.globals['_'] = lambda s, **kw: s
    app.jinja_env.globals['_l'] = lambda s, **kw: s


@app.route('/usuario/idioma', methods=['POST'])
@login_required
def trocar_idioma():
    """Atualiza preferência de idioma do usuário logado."""
    from flask import session as _sess
    lang = request.form.get('idioma', 'pt').strip().lower()
    if lang not in app.config.get('LANGUAGES', ['pt']):
        lang = 'pt'
    _sess['idioma'] = lang
    if current_user.is_authenticated:
        current_user.idioma = lang
        db.session.commit()
    return redirect(request.referrer or url_for('dashboard'))

# ========================================
# COMPRESSÃO GZIP/BROTLI
# ========================================
# Configurar compressão de respostas HTTP
compress = Compress()
compress.init_app(app)

# Configurações de compressão
app.config['COMPRESS_MIMETYPES'] = [
    'text/html',
    'text/css',
    'text/xml',
    'text/plain',
    'text/javascript',
    'application/json',
    'application/javascript',
    'application/xml',
    'application/xhtml+xml',
    'application/rss+xml',
    'application/atom+xml',
    'image/svg+xml'
]
app.config['COMPRESS_LEVEL'] = 6  # 1-9, padrão 6
app.config['COMPRESS_MIN_SIZE'] = 500  # Comprimir apenas arquivos > 500 bytes
app.config['COMPRESS_ALGORITHM'] = 'gzip'  # 'gzip' ou 'br' (brotli)

app.logger.info(f'Compressão HTTP habilitada: {app.config["COMPRESS_ALGORITHM"].upper()}')

# Configuração do Flask-Login
login_manager = LoginManager()
login_manager.login_view = 'auth.login'
login_manager.login_message = 'Por favor, faça login para acessar esta página.'
login_manager.login_message_category = 'info'
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    return get_user_by_id(user_id)

# ========================================
# SISTEMA DE AUDITORIA
# ========================================

def registrar_log(acao, entidade_tipo=None, entidade_id=None, entidade_descricao=None, detalhes=None):
    """
    Registra uma ação no sistema de auditoria

    Args:
        acao: Tipo de ação (criar, editar, excluir, login, logout, etc)
        entidade_tipo: Tipo da entidade afetada (relatorio, prova, referencia, usuario, etc)
        entidade_id: ID da entidade afetada
        entidade_descricao: Descrição da entidade para histórico
        detalhes: Informações adicionais em formato texto ou JSON
    """
    try:
        import json
        from flask_login import current_user

        # Obter informações do usuário atual
        usuario_id = current_user.id if current_user.is_authenticated else None
        usuario_nome = current_user.username if current_user.is_authenticated else 'Sistema'

        # Obter IP e User Agent
        ip_address = request.remote_addr
        user_agent = request.headers.get('User-Agent', '')[:500]

        # Converter detalhes para JSON se for dict
        if isinstance(detalhes, dict):
            detalhes = json.dumps(detalhes, ensure_ascii=False)

        # Criar registro de log
        log = AuditLog(
            usuario_id=usuario_id,
            usuario_nome=usuario_nome,
            acao=acao,
            entidade_tipo=entidade_tipo,
            entidade_id=entidade_id,
            entidade_descricao=entidade_descricao,
            detalhes=detalhes,
            ip_address=ip_address,
            user_agent=user_agent
        )

        db.session.add(log)
        db.session.commit()

        app.logger.info(f'Audit Log: {usuario_nome} {acao} {entidade_tipo}#{entidade_id}')

    except Exception as e:
        app.logger.error(f'Erro ao registrar log de auditoria: {str(e)}')
        db.session.rollback()

# Registrar Blueprints
app.register_blueprint(auth_bp)
app.register_blueprint(admin_bp)
app.register_blueprint(audit_bp, url_prefix='/auditoria')  # ✅ Habilitado - AuditLog existe

# Registrar Blueprints de Features v2
from fornecedor_bp import fornecedor_bp
from kanban_bp import kanban_bp
from checklist_bp import checklist_bp
from manuais_bp import manuais_bp
from publico_bp import publico_bp
from linx_bp import linx_bp
import linx_client
app.register_blueprint(fornecedor_bp, url_prefix='/fornecedores')
app.register_blueprint(kanban_bp, url_prefix='/kanban')
app.register_blueprint(checklist_bp, url_prefix='/admin/checklists')
app.register_blueprint(manuais_bp, url_prefix='/manuais')
app.register_blueprint(publico_bp, url_prefix='/publico')
app.register_blueprint(linx_bp, url_prefix='/linx')


@app.context_processor
def _inject_linx_flag():
    """Expõe `linx_enabled` aos templates (controla visibilidade dos botões ERP)."""
    try:
        return {'linx_enabled': linx_client.is_enabled()}
    except Exception:
        return {'linx_enabled': False}


# Registrar error handlers
register_error_handlers(app)


# ========================================
# OTIMIZAÇÕES DE PERFORMANCE
# ========================================

@app.after_request
def optimize_response(response):
    """Otimizações de performance e segurança nas respostas"""
    from datetime import datetime, timedelta

    # 1. CACHE HEADERS - Otimização de Performance
    if request.path.startswith('/static/'):
        # Cache agressivo para assets estáticos (1 ano)
        response.cache_control.public = True
        response.cache_control.max_age = 31536000  # 1 ano
        response.cache_control.immutable = True
        response.expires = datetime.now() + timedelta(days=365)
    elif request.path.startswith('/uploads/'):
        # Cache moderado para uploads (30 dias)
        response.cache_control.public = True
        response.cache_control.max_age = 2592000  # 30 dias
    elif response.content_type and 'text/html' in response.content_type:
        # Sem cache para HTML dinâmico
        response.cache_control.no_cache = True
        response.cache_control.no_store = True
        response.cache_control.must_revalidate = True
    elif response.content_type and 'application/json' in response.content_type:
        # Cache curto para API JSON (5 minutos)
        response.cache_control.public = True
        response.cache_control.max_age = 300

    # 2. COMPRESSION HINTS
    if 'Content-Type' in response.headers:
        content_type = response.headers['Content-Type']
        compressible = ['text/', 'application/json', 'application/javascript',
                       'application/xml', 'image/svg+xml']
        if any(ct in content_type for ct in compressible):
            response.vary = 'Accept-Encoding'

    return response


def gerar_e_salvar_pdf(relatorio_id, evento="CRIADO"):
    """
    Geração de PDF desabilitada temporariamente.
    Para habilitar: pip install xhtml2pdf e descomentar import acima
    """
    print(f"PDF não gerado (xhtml2pdf não instalado) para relatório ID: {relatorio_id} (Evento: {evento})")
    return True  # Retornar True para não quebrar o fluxo

    # CÓDIGO COMENTADO - Descomente após instalar xhtml2pdf
    # print(f"Iniciando geração de PDF para o relatório ID: {relatorio_id} (Evento: {evento})")
    #
    # relatorio = Relatorio.query.get(relatorio_id)
    # if not relatorio:
    #     print(f"!!! ERRO: Relatório com ID {relatorio_id} não encontrado para gerar PDF.")
    #     return False
    #
    # # Preparar dados para o template (mantendo estrutura compatível)
    # referencias_completas = []
    # for ref in relatorio.referencias:
    #     ref_dict = {c.name: getattr(ref, c.name) for c in ref.__table__.columns}
    #
    #     provas_completas = []
    #     for prova in ref.provas:
    #         prova_dict = {c.name: getattr(prova, c.name) for c in prova.__table__.columns}
    #
    #         prova_dict['fotos'] = {}
    #         for foto in prova.fotos:
    #             contexto = foto.contexto
    #             if contexto not in prova_dict['fotos']:
    #                 prova_dict['fotos'][contexto] = []
    #             prova_dict['fotos'][contexto].append({c.name: getattr(foto, c.name) for c in foto.__table__.columns})
    #
    #         provas_completas.append(prova_dict)
    #
    #     ref_dict['provas'] = provas_completas
    #     referencias_completas.append(ref_dict)
    #
    # assunto_email = f"RELATÓRIO DE PROVA PEÇA PILOTO {evento}! {relatorio.descricao_geral} COLEÇÃO {relatorio.colecao}"
    # nome_ficheiro_pdf = f"{secure_filename(assunto_email)}.pdf"
    # caminho_ficheiro_pdf = os.path.join(app.config['PDF_FOLDER'], nome_ficheiro_pdf)
    #
    # html_renderizado = render_template('relatorio_pdf.html',
    #                                    relatorio=relatorio,
    #                                    referencias=referencias_completas)
    #
    # def link_callback(uri, rel):
    #     if uri.startswith(url_for('serve_upload', filename='')):
    #         path_relativo = uri[len(url_for('serve_upload', filename='')):]
    #         caminho_final = os.path.join(app.config['UPLOAD_FOLDER'], path_relativo)
    #         return caminho_final
    #     return uri
    #
    # try:
    #     with open(caminho_ficheiro_pdf, "w+b") as pdf_file:
    #         pisa_status = pisa.CreatePDF(
    #             html_renderizado.encode('utf-8'),
    #             dest=pdf_file,
    #             encoding='utf-8',
    #             link_callback=link_callback
    #         )
    #
    #     if not pisa_status.err:
    #         print(f"\n--- PDF Gerado: {caminho_ficheiro_pdf} ---\n")
    #         return True
    #     else:
    #         print(f"!!! ERRO AO GERAR PDF: {pisa_status.err}")
    #         return False
    # except Exception as e:
    #     print(f"!!! EXCEÇÃO AO GERAR PDF: {e}")
    #     return False

@app.route('/')
@login_required
def dashboard():
    # Paginação
    page = request.args.get('page', 1, type=int)
    per_page = 20

    # Filtro por referência
    filtro_referencia = request.args.get('referencia', '').strip()

    # Obter relatórios paginados (com filtro opcional por referência)
    query = Relatorio.query
    if filtro_referencia:
        query = query.join(Referencia).filter(
            Referencia.numero_ref.ilike(f'%{filtro_referencia}%')
        ).distinct()
    pagination = query.order_by(desc(Relatorio.created_at)).paginate(
        page=page, per_page=per_page, error_out=False
    )
    relatorios = pagination.items
    relatorios_com_status = []

    for relatorio in relatorios:
        relatorio_dict = {c.name: getattr(relatorio, c.name) for c in relatorio.__table__.columns}

        # Obter última prova de qualquer referência deste relatório
        ultima_prova = Prova.query.join(Referencia).filter(Referencia.relatorio_id == relatorio.id).order_by(desc(Prova.numero_prova)).first()

        relatorio_dict['status_atual'] = ultima_prova.status if ultima_prova else 'Novo'

        # Buscar primeira foto do relatório
        # Ordem de prioridade: imagem_produto do relatório > primeira foto de qualquer prova
        primeira_foto = None

        if relatorio.imagem_produto:
            primeira_foto = relatorio.imagem_produto
        else:
            # Buscar primeira foto de qualquer prova deste relatório
            foto = Foto.query.join(Prova).join(Referencia).filter(
                Referencia.relatorio_id == relatorio.id
            ).order_by(Foto.id).first()

            if foto:
                primeira_foto = foto.file_path

        relatorio_dict['primeira_foto'] = primeira_foto

        # Buscar informações das referências
        referencias = Referencia.query.filter_by(relatorio_id=relatorio.id).all()
        relatorio_dict['total_referencias'] = len(referencias)

        # Categoria (primeira categoria encontrada)
        categorias = [r.tipo_categoria for r in referencias if r.tipo_categoria]
        relatorio_dict['categoria'] = categorias[0] if categorias else None

        # Fornecedores únicos
        fornecedores = list(set([r.fornecedor for r in referencias if r.fornecedor]))
        relatorio_dict['fornecedores'] = ', '.join(fornecedores) if fornecedores else None

        # F2: Prazo / Deadline badge
        if relatorio.data_limite:
            try:
                from datetime import date
                data_limite = date.fromisoformat(relatorio.data_limite)
                hoje = date.today()
                dias_restantes = (data_limite - hoje).days
                if dias_restantes < 0:
                    relatorio_dict['prazo_status'] = 'vencido'
                    relatorio_dict['prazo_label'] = f'Vencido há {abs(dias_restantes)} dia(s)'
                elif dias_restantes <= 7:
                    relatorio_dict['prazo_status'] = 'proximo'
                    relatorio_dict['prazo_label'] = f'{dias_restantes} dia(s) restante(s)'
                else:
                    relatorio_dict['prazo_status'] = 'ok'
                    relatorio_dict['prazo_label'] = f'{dias_restantes} dias'
                relatorio_dict['data_limite'] = relatorio.data_limite
            except (ValueError, TypeError):
                relatorio_dict['prazo_status'] = None
                relatorio_dict['data_limite'] = relatorio.data_limite
        else:
            relatorio_dict['prazo_status'] = None

        # Data de criação formatada
        if relatorio.created_at:
            relatorio_dict['data_criacao'] = relatorio.created_at.strftime('%d/%m/%Y')
        else:
            relatorio_dict['data_criacao'] = 'N/A'

        # Total de provas
        total_provas = Prova.query.join(Referencia).filter(Referencia.relatorio_id == relatorio.id).count()
        relatorio_dict['total_provas'] = total_provas

        relatorios_com_status.append(relatorio_dict)

    # ESTATÍSTICAS E INSIGHTS
    total_relatorios = Relatorio.query.count()
    total_referencias = Referencia.query.count()
    total_provas = Prova.query.count()

    # Provas por status - Valores em MAIÚSCULAS após padronização
    provas_aprovadas = Prova.query.filter_by(status='APROVADA').count()
    provas_reprovadas = Prova.query.filter_by(status='REPROVADA').count()
    provas_em_andamento = Prova.query.filter_by(status='EM ANDAMENTO').count()
    provas_comite = Prova.query.filter_by(status='COMITÊ').count()

    # Taxa de aprovação
    taxa_aprovacao = round((provas_aprovadas / total_provas * 100) if total_provas > 0 else 0, 1)

    # Referências por categoria
    refs_por_categoria = db.session.query(
        Referencia.tipo_categoria,
        db.func.count(Referencia.id)
    ).group_by(Referencia.tipo_categoria).all()

    categorias_stats = {cat: count for cat, count in refs_por_categoria}

    # Relatórios recentes (últimos 30 dias)
    from datetime import datetime, timedelta
    trinta_dias_atras = datetime.utcnow() - timedelta(days=30)
    relatorios_recentes = Relatorio.query.filter(Relatorio.created_at >= trinta_dias_atras).count()

    # Provas com retrabalho (número da prova > 1)
    provas_retrabalho = Prova.query.filter(Prova.numero_prova > 1).count()
    taxa_retrabalho = round((provas_retrabalho / total_provas * 100) if total_provas > 0 else 0, 1)

    # Insights
    insights = []

    if taxa_aprovacao >= 80:
        insights.append({
            'tipo': 'success',
            'icone': 'bi-trophy',
            'titulo': 'Excelente Performance!',
            'mensagem': f'Taxa de aprovação de {taxa_aprovacao}% - acima da meta de 80%'
        })
    elif taxa_aprovacao >= 60:
        insights.append({
            'tipo': 'warning',
            'icone': 'bi-exclamation-triangle',
            'titulo': 'Performance Moderada',
            'mensagem': f'Taxa de aprovação de {taxa_aprovacao}% - pode melhorar'
        })
    else:
        insights.append({
            'tipo': 'danger',
            'icone': 'bi-exclamation-circle',
            'titulo': 'Atenção Necessária',
            'mensagem': f'Taxa de aprovação baixa: {taxa_aprovacao}% - revisar processos'
        })

    if taxa_retrabalho > 30:
        insights.append({
            'tipo': 'warning',
            'icone': 'bi-arrow-repeat',
            'titulo': 'Alto Retrabalho',
            'mensagem': f'{taxa_retrabalho}% das provas precisaram de retrabalho'
        })

    if relatorios_recentes > 5:
        insights.append({
            'tipo': 'info',
            'icone': 'bi-calendar-check',
            'titulo': 'Alta Produtividade',
            'mensagem': f'{relatorios_recentes} relatórios criados nos últimos 30 dias'
        })

    if provas_em_andamento > 10:
        insights.append({
            'tipo': 'info',
            'icone': 'bi-hourglass-split',
            'titulo': 'Provas Pendentes',
            'mensagem': f'{provas_em_andamento} provas aguardando finalização'
        })

    # Média de provas por referência
    media_provas_por_referencia = round((total_provas / total_referencias) if total_referencias > 0 else 0, 1)

    stats = {
        'total_relatorios': total_relatorios,
        'total_referencias': total_referencias,
        'total_provas': total_provas,
        'provas_aprovadas': provas_aprovadas,
        'provas_reprovadas': provas_reprovadas,
        'provas_em_andamento': provas_em_andamento,
        'provas_comite': provas_comite,
        'taxa_aprovacao': taxa_aprovacao,
        'taxa_retrabalho': taxa_retrabalho,
        'media_provas_por_referencia': media_provas_por_referencia,
        'categorias': categorias_stats,
        'relatorios_recentes': relatorios_recentes,
        'insights': insights
    }

    # F9: Carregar preferências do usuário
    widget_prefs = {}
    try:
        pref = PreferenciaUsuario.query.filter_by(usuario_id=current_user.id).first()
        if pref:
            widget_prefs = pref.get_config()
    except Exception:
        pass

    return render_template('dashboard.html', relatorios=relatorios_com_status, stats=stats, pagination=pagination, widget_prefs=widget_prefs)

@app.route('/favicon.ico')
def favicon():
    """Serve favicon from static folder"""
    return send_from_directory(
        os.path.join(app.root_path, 'static'),
        'favicon.ico',
        mimetype='image/vnd.microsoft.icon'
    )

@app.route('/uploads/<path:filename>')
@login_required
def serve_upload(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/relatorio/<int:id>')
@login_required
def detalhes_relatorio(id):
    relatorio = Relatorio.query.get_or_404(id)
    
    # Preparar estrutura para template
    referencias_completas = []
    for ref in relatorio.referencias:
        ref_dict = {c.name: getattr(ref, c.name) for c in ref.__table__.columns}
        
        provas_completas = []
        # Ordenar provas por número
        provas_ordenadas = sorted(ref.provas, key=lambda x: x.numero_prova)
        
        for prova in provas_ordenadas:
            prova_dict = {c.name: getattr(prova, c.name) for c in prova.__table__.columns}

            prova_dict['fotos'] = {}
            for foto in prova.fotos:
                contexto = foto.contexto
                if contexto not in prova_dict['fotos']:
                    prova_dict['fotos'][contexto] = []
                prova_dict['fotos'][contexto].append({c.name: getattr(foto, c.name) for c in foto.__table__.columns})

            # Respostas estruturadas (ChecklistTemplate dinâmico) agrupadas por categoria
            prova_dict['respostas_dinamicas'] = {'qualidade': [], 'estilo': [], 'modelagem': []}
            for resposta in (prova.respostas or []):
                if resposta.template and resposta.template.categoria in prova_dict['respostas_dinamicas']:
                    prova_dict['respostas_dinamicas'][resposta.template.categoria].append({
                        'item': resposta.item,
                        'conforme': resposta.conforme,
                        'observacao': resposta.observacao,
                    })

            provas_completas.append(prova_dict)
        
        ref_dict['provas'] = provas_completas
        referencias_completas.append(ref_dict)

    # Calcular prazo_status para exibição visual
    prazo_status = None
    if relatorio.data_limite:
        from datetime import date
        try:
            prazo_date = date.fromisoformat(relatorio.data_limite)
            hoje = date.today()
            diff = (prazo_date - hoje).days
            if diff < 0:
                prazo_status = 'vencido'
            elif diff <= 7:
                prazo_status = 'proximo'
            else:
                prazo_status = 'ok'
        except (ValueError, TypeError):
            pass

    # Links públicos ativos deste relatório
    from models import LinkPublico
    links_publicos = (
        LinkPublico.query
        .filter_by(relatorio_id=relatorio.id, is_active=True)
        .order_by(LinkPublico.created_at.desc())
        .all()
    )

    return render_template(
        'detalhes_relatorio.html',
        relatorio=relatorio,
        referencias=referencias_completas,
        prazo_status=prazo_status,
        links_publicos=links_publicos,
    )

def _salvar_checklist_dinamico_para_prova(prova_id, categoria, itens_marcados):
    """Substitui as ChecklistResposta do prova/categoria pelos itens marcados.

    - Busca o ChecklistTemplate ativo da categoria
    - Remove respostas existentes vinculadas àquele template
    - Cria novos registros para cada item marcado
    Se não há template ativo, no-op (não bloqueia o save).
    """
    template = (
        ChecklistTemplate.query
        .filter_by(categoria=categoria, is_active=True)
        .order_by(ChecklistTemplate.updated_at.desc(), ChecklistTemplate.id.desc())
        .first()
    )
    if not template:
        return

    # Remover respostas anteriores deste prova+template
    ChecklistResposta.query.filter_by(prova_id=prova_id, template_id=template.id).delete()

    for item in itens_marcados or []:
        item_str = (item or '').strip()
        if not item_str:
            continue
        db.session.add(ChecklistResposta(
            prova_id=prova_id,
            template_id=template.id,
            item=item_str,
            conforme=True,
        ))


def _construir_payload_pdf(relatorio):
    """Monta a lista de referencias_completas com fotos em base64 para o template do PDF.
    Reutilizado por /relatorio/<id>/pdf (login) e /publico/<token>/pdf (sem login)."""
    import base64
    import mimetypes

    referencias_completas = []
    for ref in relatorio.referencias:
        ref_dict = {c.name: getattr(ref, c.name) for c in ref.__table__.columns}

        provas_completas = []
        provas_ordenadas = sorted(ref.provas, key=lambda x: x.numero_prova)

        for prova in provas_ordenadas:
            prova_dict = {c.name: getattr(prova, c.name) for c in prova.__table__.columns}

            prova_dict['fotos'] = {}
            for foto in prova.fotos:
                contexto = foto.contexto
                if contexto not in prova_dict['fotos']:
                    prova_dict['fotos'][contexto] = []
                foto_dict = {c.name: getattr(foto, c.name) for c in foto.__table__.columns}
                caminho_absoluto = os.path.join(app.config['UPLOAD_FOLDER'], foto.file_path)
                if os.path.exists(caminho_absoluto):
                    try:
                        mime_type = mimetypes.guess_type(caminho_absoluto)[0] or 'image/jpeg'
                        with open(caminho_absoluto, 'rb') as img_file:
                            img_data = base64.b64encode(img_file.read()).decode('utf-8')
                        foto_dict['base64'] = f'data:{mime_type};base64,{img_data}'
                    except Exception as e:
                        app.logger.error(f'Erro ao converter foto {foto.file_path} para base64: {e}')
                        foto_dict['base64'] = ''
                else:
                    foto_dict['base64'] = ''
                prova_dict['fotos'][contexto].append(foto_dict)

            provas_completas.append(prova_dict)

        ref_dict['provas'] = provas_completas
        referencias_completas.append(ref_dict)
    return referencias_completas


def _gerar_pdf_relatorio_response(relatorio, inline=True):
    """Gera o PDF do relatório e devolve uma Response Flask."""
    from datetime import datetime as _dt
    referencias_completas = _construir_payload_pdf(relatorio)
    html_string = render_template(
        'relatorio_pdf.html',
        relatorio=relatorio,
        referencias=referencias_completas,
        now=_dt.now,
    )
    base_url = 'file://' + app.config.get('UPLOAD_FOLDER', '') + '/'
    pdf = HTML(string=html_string, base_url=base_url).write_pdf()
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    disposition = 'inline' if inline else 'attachment'
    response.headers['Content-Disposition'] = (
        f"{disposition}; filename=relatorio_{relatorio.id}_{secure_filename(relatorio.descricao_geral)}.pdf"
    )
    return response


@app.route('/relatorio/<int:id>/pdf')
@login_required
def relatorio_pdf(id):
    """Gera e retorna o PDF do relatório (versão autenticada)."""
    relatorio = Relatorio.query.get_or_404(id)
    return _gerar_pdf_relatorio_response(relatorio, inline=True)

@app.route('/relatorio/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_relatorio(id):
    relatorio = Relatorio.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # 1. Atualiza as informações gerais do relatório
            relatorio.colecao = request.form.get('colecao')
            relatorio.descricao_geral = request.form.get('descricao_geral')
            relatorio.linha = request.form.get('linha')

            # F2: Atualizar data_limite
            relatorio.data_limite = request.form.get('data_limite')

            # Atualiza PPT se um novo arquivo foi enviado
            ppt_file = request.files.get('ppt')
            if ppt_file and ppt_file.filename:
                # F8: Salvar versão anterior
                if relatorio.ppt_path:
                    salvar_versao_arquivo('relatorio', relatorio.id, 'ppt_path', relatorio.ppt_path)
                # Salvar novo PPT
                ppt_filename = save_file(ppt_file)
                if ppt_filename:
                    relatorio.ppt_path = ppt_filename

            # Atualiza Imagem do Produto se um novo arquivo foi enviado
            imagem_file = request.files.get('imagem_produto')
            if imagem_file and imagem_file.filename:
                # F8: Salvar versão anterior
                if relatorio.imagem_produto:
                    salvar_versao_arquivo('relatorio', relatorio.id, 'imagem_produto', relatorio.imagem_produto)
                imagem_filename = save_file(imagem_file)
                if imagem_filename:
                    relatorio.imagem_produto = imagem_filename

            # Atualiza Ficha Técnica se um novo arquivo foi enviado
            ficha_file = request.files.get('ficha_tecnica')
            if ficha_file and ficha_file.filename:
                # F8: Salvar versão anterior
                if relatorio.ficha_tecnica:
                    from utils import delete_file
                    delete_file(relatorio.ficha_tecnica)
                ficha_filename = save_file(ficha_file)
                if ficha_filename:
                    relatorio.ficha_tecnica = ficha_filename
            
            # 2. Itera sobre todos os tipos possíveis
            for tipo in ['baby', 'kids', 'teen', 'adulto']:
                ref_numero = request.form.get(f'ref_{tipo}')
                if not ref_numero:
                    continue

                # Verifica se já existe uma referência deste tipo
                ref_existente = Referencia.query.filter_by(relatorio_id=id, tipo_categoria=tipo).first()

                if ref_existente:
                    # --- ATUALIZAÇÃO ---
                    ref_existente.numero_ref = ref_numero
                    # Atualiza dados da referência (busca por ID específico ou por tipo)
                    ref_existente.origem = request.form.get(f'origem_ref_{ref_existente.id}') or request.form.get(f'origem_{tipo}')
                    ref_existente.fornecedor = request.form.get(f'fornecedor_ref_{ref_existente.id}') or request.form.get(f'fornecedor_{tipo}')
                    ref_existente.materia_prima = request.form.get(f'materia_prima_ref_{ref_existente.id}') or request.form.get(f'materia_prima_{tipo}')
                    ref_existente.composicao = request.form.get(f'composicao_ref_{ref_existente.id}') or request.form.get(f'composicao_{tipo}')
                    ref_existente.gramatura = request.form.get(f'gramatura_ref_{ref_existente.id}') or request.form.get(f'gramatura_{tipo}')
                    ref_existente.aviamentos = request.form.get(f'aviamentos_ref_{ref_existente.id}') or request.form.get(f'aviamentos_{tipo}')
                    
                    # Atualiza provas existentes
                    provas_existentes_ids = request.form.getlist(f'prova_id_{tipo}')
                    for prova_id_str in provas_existentes_ids:
                        prova_id = int(prova_id_str)
                        prova = Prova.query.get(prova_id)
                        if prova:
                            # Atualizar tabela de medidas se novo arquivo enviado
                            tabela_file = request.files.get(f'tabela_medidas_{prova_id}')
                            if tabela_file and tabela_file.filename:
                                tabela_filename = save_file(tabela_file)
                                if tabela_filename:
                                    prova.tabela_medidas_path = tabela_filename

                            prova.data_recebimento = request.form.get(f'data_recebimento_{prova_id}')
                            prova.tamanhos_recebidos = ", ".join(request.form.getlist(f'tamanhos_recebidos_{prova_id}'))
                            prova.info_medidas = request.form.get(f'info_medidas_{prova_id}')
                            prova.data_prova = request.form.get(f'data_prova_{prova_id}')
                            prova.time_qualidade = request.form.get(f'time_qualidade_{prova_id}')
                            prova.checklist_qualidade = ", ".join(request.form.getlist(f'checklist_qualidade_{prova_id}'))
                            prova.comentarios_qualidade = request.form.get(f'comentarios_qualidade_{prova_id}')
                            prova.obs_qualidade = request.form.get(f'obs_qualidade_{prova_id}')
                            prova.time_estilo = request.form.get(f'time_estilo_{prova_id}')
                            prova.checklist_estilo = ", ".join(request.form.getlist(f'checklist_estilo_{prova_id}'))
                            prova.comentarios_estilo = request.form.get(f'comentarios_estilo_{prova_id}')
                            prova.obs_estilo = request.form.get(f'obs_estilo_{prova_id}')
                            prova.time_modelagem = request.form.get(f'time_modelagem_{prova_id}')
                            prova.checklist_modelagem = ", ".join(request.form.getlist(f'checklist_modelagem_{prova_id}'))
                            prova.comentarios_modelagem = request.form.get(f'comentarios_modelagem_{prova_id}')
                            prova.obs_modelagem = request.form.get(f'obs_modelagem_{prova_id}')

                            # Persistir respostas do checklist dinâmico (ChecklistTemplate ativo por categoria)
                            for cat in ('qualidade', 'estilo', 'modelagem'):
                                _salvar_checklist_dinamico_para_prova(prova.id, cat,
                                    request.form.getlist(f'checklist_dinamico_{cat}_{prova_id}'))

                            prova.data_lacre = request.form.get(f'data_lacre_{prova_id}')
                            prova.numero_lacre = request.form.get(f'numero_lacre_{prova_id}')
                            prova.info_adicionais = request.form.get(f'info_adicionais_{prova_id}')
                            
                            # Adicionar fotos
                            campos_fotos = ['desenho', 'qualidade', 'estilo', 'modelagem']
                            for contexto in campos_fotos:
                                for file in request.files.getlist(f'fotos_{contexto}_{prova_id}'):
                                    filename = save_file(file)
                                    if filename:
                                        foto = Foto(prova_id=prova.id, contexto=contexto, file_path=filename)
                                        db.session.add(foto)

                            # Upload direto de fotos amostra/prova_modelo (sem tamanho)
                            for contexto in ['amostra', 'prova_modelo']:
                                for file in request.files.getlist(f'fotos_{contexto}_{prova_id}'):
                                    filename = save_file(file)
                                    if filename:
                                        foto = Foto(prova_id=prova.id, contexto=contexto, file_path=filename)
                                        db.session.add(foto)

                            # Upload por tamanho (opcional)
                            for tamanho in request.form.getlist(f'tamanhos_recebidos_{prova_id}'):
                                for contexto in ['amostra', 'prova_modelo']:
                                    for file in request.files.getlist(f'fotos_{contexto}_{prova_id}_{tamanho.replace(" ", "")}'):
                                        filename = save_file(file)
                                        if filename:
                                            foto = Foto(prova_id=prova.id, contexto=contexto, tamanho=tamanho, file_path=filename)
                                            db.session.add(foto)

                else:
                    # --- CRIAÇÃO ---
                    nova_ref = Referencia(
                        relatorio_id=id,
                        tipo_categoria=tipo,
                        numero_ref=ref_numero,
                        origem=request.form.get(f'origem_{tipo}'),
                        fornecedor=request.form.get(f'fornecedor_{tipo}'),
                        materia_prima=request.form.get(f'materia_prima_{tipo}'),
                        composicao=request.form.get(f'composicao_{tipo}'),
                        gramatura=request.form.get(f'gramatura_{tipo}'),
                        aviamentos=request.form.get(f'aviamentos_{tipo}')
                    )
                    db.session.add(nova_ref)
                    db.session.flush() # Para obter o ID

                    tabela_medidas_filename = save_file(request.files.get(f'tabela_medidas_{tipo}'))

                    nova_prova = Prova(
                        referencia_id=nova_ref.id,
                        numero_prova=1,
                        tabela_medidas_path=tabela_medidas_filename,
                        data_recebimento=request.form.get(f'data_recebimento_{tipo}'),
                        tamanhos_recebidos=", ".join(request.form.getlist(f'tamanhos_recebidos_{tipo}')),
                        info_medidas=request.form.get(f'info_medidas_{tipo}'),
                        data_prova=request.form.get(f'data_prova_{tipo}'),
                        time_qualidade=request.form.get(f'time_qualidade_{tipo}'),
                        checklist_qualidade=", ".join(request.form.getlist(f'checklist_qualidade_{tipo}')),
                        comentarios_qualidade=request.form.get(f'comentarios_qualidade_{tipo}'),
                        obs_qualidade=request.form.get(f'obs_qualidade_{tipo}'),
                        time_estilo=request.form.get(f'time_estilo_{tipo}'),
                        checklist_estilo=", ".join(request.form.getlist(f'checklist_estilo_{tipo}')),
                        comentarios_estilo=request.form.get(f'comentarios_estilo_{tipo}'),
                        obs_estilo=request.form.get(f'obs_estilo_{tipo}'),
                        time_modelagem=request.form.get(f'time_modelagem_{tipo}'),
                        checklist_modelagem=", ".join(request.form.getlist(f'checklist_modelagem_{tipo}')),
                        comentarios_modelagem=request.form.get(f'comentarios_modelagem_{tipo}'),
                        obs_modelagem=request.form.get(f'obs_modelagem_{tipo}'),
                        data_lacre=request.form.get(f'data_lacre_{tipo}'),
                        numero_lacre=request.form.get(f'numero_lacre_{tipo}'),
                        info_adicionais=request.form.get(f'info_adicionais_{tipo}')
                    )
                    db.session.add(nova_prova)
                    db.session.flush()

                    campos_fotos = ['desenho', 'qualidade', 'estilo', 'modelagem']
                    for contexto in campos_fotos:
                        for file in request.files.getlist(f'fotos_{contexto}_{tipo}'):
                            filename = save_file(file)
                            if filename:
                                foto = Foto(prova_id=nova_prova.id, contexto=contexto, file_path=filename)
                                db.session.add(foto)

                    # Upload direto de fotos amostra/prova_modelo (sem tamanho)
                    for contexto in ['amostra', 'prova_modelo']:
                        for file in request.files.getlist(f'fotos_{contexto}_{tipo}'):
                            filename = save_file(file)
                            if filename:
                                foto = Foto(prova_id=nova_prova.id, contexto=contexto, file_path=filename)
                                db.session.add(foto)

                    # Upload por tamanho (opcional)
                    for tamanho in request.form.getlist(f'tamanhos_recebidos_{tipo}'):
                        for contexto in ['amostra', 'prova_modelo']:
                            for file in request.files.getlist(f'fotos_{contexto}_{tipo}_{tamanho.replace(" ", "")}'):
                                filename = save_file(file)
                                if filename:
                                    foto = Foto(prova_id=nova_prova.id, contexto=contexto, tamanho=tamanho, file_path=filename)
                                    db.session.add(foto)

            db.session.commit()
            flash("Relatório atualizado com sucesso!", "success")
            gerar_e_salvar_pdf(id, evento="ATUALIZADO")
            return redirect(url_for('detalhes_relatorio', id=id))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao atualizar o relatório: {e}", "error")
            return redirect(url_for('editar_relatorio', id=id))
    
    # GET
    referencias_por_tipo = {}
    for ref in relatorio.referencias:
        ref_dict = {c.name: getattr(ref, c.name) for c in ref.__table__.columns}

        provas_completas = []
        provas_ordenadas = sorted(ref.provas, key=lambda x: x.numero_prova)
        for prova in provas_ordenadas:
            prova_dict = {c.name: getattr(prova, c.name) for c in prova.__table__.columns}
            prova_dict['tamanhos_lista'] = [t.strip() for t in (prova.tamanhos_recebidos or '').split(',') if t.strip()]

            prova_dict['fotos'] = {}
            for foto in prova.fotos:
                contexto = foto.contexto
                if contexto not in prova_dict['fotos']:
                    prova_dict['fotos'][contexto] = []
                prova_dict['fotos'][contexto].append({c.name: getattr(foto, c.name) for c in foto.__table__.columns})
            provas_completas.append(prova_dict)

        ref_dict['provas'] = provas_completas
        tipo_lower = ref.tipo_categoria.lower() if ref.tipo_categoria else ""
        referencias_por_tipo[tipo_lower] = ref_dict

    return render_template('editar_relatorio.html', relatorio=relatorio, referencias_por_tipo=referencias_por_tipo)


@app.route('/prova/atualizar_status', methods=['POST'])
@login_required
def atualizar_status():
    prova_id = request.form.get('prova_id', type=int)
    novo_status = request.form.get('novo_status')
    motivo = request.form.get('motivo')

    if not prova_id:
        flash("ID da prova inválido.", "error")
        return redirect(url_for('dashboard'))

    try:
        prova = Prova.query.get(prova_id)
        if prova:
            prova.status = novo_status
            prova.motivo_ultima_alteracao = motivo
            db.session.commit()
            
            flash(f"Status da prova atualizado para '{novo_status}' com sucesso!", "success")
            
            # Encontrar relatório para gerar PDF
            relatorio_id = prova.referencia.relatorio_id
            gerar_e_salvar_pdf(relatorio_id, evento="ATUALIZADO")
            return redirect(url_for('detalhes_relatorio', id=relatorio_id))
            
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao atualizar o status: {e}", "error")
        return redirect(url_for('dashboard'))
    
    return redirect(url_for('dashboard'))

@app.route('/novo', methods=['GET', 'POST'])
@login_required
def novo_relatorio():
    if request.method == 'POST':
        try:
            ppt_filename = save_file(request.files.get('ppt'))
            imagem_produto_filename = save_file(request.files.get('imagem_produto'))
            ficha_tecnica_filename = save_file(request.files.get('ficha_tecnica'))

            novo_relatorio = Relatorio(
                codigo=gerar_codigo_relatorio(),
                descricao_geral=request.form.get('descricao_geral'),
                colecao=request.form.get('colecao'),
                linha=request.form.get('linha'),
                data_limite=request.form.get('data_limite'),
                ppt_path=ppt_filename,
                imagem_produto=imagem_produto_filename,
                ficha_tecnica=ficha_tecnica_filename,
                created_by=current_user.id if current_user.is_authenticated else None
            )
            db.session.add(novo_relatorio)
            db.session.flush()
            
            for tipo in ['baby', 'kids', 'teen', 'adulto']:
                if request.form.get(f'ref_{tipo}'):
                    nova_ref = Referencia(
                        relatorio_id=novo_relatorio.id,
                        tipo_categoria=tipo,
                        numero_ref=request.form.get(f'ref_{tipo}'),
                        origem=request.form.get(f'origem_{tipo}'),
                        fornecedor=request.form.get(f'fornecedor_{tipo}'),
                        materia_prima=request.form.get(f'materia_prima_{tipo}'),
                        composicao=request.form.get(f'composicao_{tipo}'),
                        gramatura=request.form.get(f'gramatura_{tipo}'),
                        aviamentos=request.form.get(f'aviamentos_{tipo}')
                    )
                    db.session.add(nova_ref)
                    db.session.flush()

                    tabela_medidas_filename = save_file(request.files.get(f'tabela_medidas_{tipo}'))

                    nova_prova = Prova(
                        referencia_id=nova_ref.id,
                        numero_prova=1,
                        tabela_medidas_path=tabela_medidas_filename,
                        data_recebimento=request.form.get(f'data_recebimento_{tipo}'),
                        tamanhos_recebidos=", ".join(request.form.getlist(f'tamanhos_recebidos_{tipo}')),
                        info_medidas=request.form.get(f'info_medidas_{tipo}'),
                        data_prova=request.form.get(f'data_prova_{tipo}'),
                        time_qualidade=request.form.get(f'time_qualidade_{tipo}'),
                        checklist_qualidade=", ".join(request.form.getlist(f'checklist_qualidade_{tipo}')),
                        comentarios_qualidade=request.form.get(f'comentarios_qualidade_{tipo}'),
                        obs_qualidade=request.form.get(f'obs_qualidade_{tipo}'),
                        time_estilo=request.form.get(f'time_estilo_{tipo}'),
                        checklist_estilo=", ".join(request.form.getlist(f'checklist_estilo_{tipo}')),
                        comentarios_estilo=request.form.get(f'comentarios_estilo_{tipo}'),
                        obs_estilo=request.form.get(f'obs_estilo_{tipo}'),
                        time_modelagem=request.form.get(f'time_modelagem_{tipo}'),
                        checklist_modelagem=", ".join(request.form.getlist(f'checklist_modelagem_{tipo}')),
                        comentarios_modelagem=request.form.get(f'comentarios_modelagem_{tipo}'),
                        obs_modelagem=request.form.get(f'obs_modelagem_{tipo}'),
                        data_lacre=request.form.get(f'data_lacre_{tipo}'),
                        numero_lacre=request.form.get(f'numero_lacre_{tipo}'),
                        info_adicionais=request.form.get(f'info_adicionais_{tipo}')
                    )
                    db.session.add(nova_prova)
                    db.session.flush()

                    campos_fotos = ['desenho', 'qualidade', 'estilo', 'modelagem']
                    for contexto in campos_fotos:
                        for file in request.files.getlist(f'fotos_{contexto}_{tipo}'):
                            filename = save_file(file)
                            if filename:
                                foto = Foto(prova_id=nova_prova.id, contexto=contexto, file_path=filename)
                                db.session.add(foto)

                    # Upload direto de fotos amostra/prova_modelo (sem tamanho)
                    for contexto in ['amostra', 'prova_modelo']:
                        for file in request.files.getlist(f'fotos_{contexto}_{tipo}'):
                            filename = save_file(file)
                            if filename:
                                foto = Foto(prova_id=nova_prova.id, contexto=contexto, file_path=filename)
                                db.session.add(foto)

                    # Upload por tamanho (opcional)
                    for tamanho in request.form.getlist(f'tamanhos_recebidos_{tipo}'):
                        for contexto in ['amostra', 'prova_modelo']:
                            for file in request.files.getlist(f'fotos_{contexto}_{tipo}_{tamanho.replace(" ", "")}'):
                                filename = save_file(file)
                                if filename:
                                    foto = Foto(prova_id=nova_prova.id, contexto=contexto, tamanho=tamanho, file_path=filename)
                                    db.session.add(foto)

            db.session.commit()
            flash("Relatório criado com sucesso!", "success")
            
            gerar_e_salvar_pdf(novo_relatorio.id, evento="CRIADO")
            
            return redirect(url_for('dashboard'))

        except Exception as e:
            db.session.rollback()
            print(f"!!! ERRO AO SALVAR NO BANCO DE DADOS: {e}")
            flash(f"Ocorreu um erro ao salvar o relatório: {e}. Por favor, verifique os campos e tente novamente.", "error")
            return render_template('novo_relatorio.html', form_data=request.form)

    return render_template('novo_relatorio.html', form_data=MultiDict())

@app.route('/relatorio/<int:id>/excluir', methods=['POST'])
@login_required
def excluir_relatorio(id):
    """Exclui um relatório e todos os seus arquivos associados"""
    relatorio = Relatorio.query.get_or_404(id)

    try:
        # Coletar todos os arquivos para excluir
        arquivos_para_excluir = []

        # PPT do relatório
        if relatorio.ppt_path:
            arquivos_para_excluir.append(relatorio.ppt_path)

        # Arquivos de referências e provas
        for ref in relatorio.referencias:
            for prova in ref.provas:
                # Tabela de medidas
                if prova.tabela_medidas_path:
                    arquivos_para_excluir.append(prova.tabela_medidas_path)
                # Fotos
                for foto in prova.fotos:
                    if foto.file_path:
                        arquivos_para_excluir.append(foto.file_path)

        # Excluir do banco (cascade delete cuida das referências, provas e fotos)
        descricao = relatorio.descricao_geral
        colecao = relatorio.colecao
        db.session.delete(relatorio)
        db.session.commit()

        # Registrar log de auditoria
        registrar_log(
            acao='excluir',
            entidade_tipo='relatorio',
            entidade_id=id,
            entidade_descricao=f'{descricao} - {colecao}',
            detalhes=f'Excluído {len(arquivos_para_excluir)} arquivo(s) associado(s)'
        )

        # Excluir arquivos físicos
        from utils import delete_file
        for arquivo in arquivos_para_excluir:
            delete_file(arquivo)

        flash(f"Relatório '{descricao}' excluído com sucesso!", "success")
        return redirect(url_for('dashboard'))

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir o relatório: {e}", "error")
        return redirect(url_for('detalhes_relatorio', id=id))


@app.route('/relatorio/<int:id>/excluir_arquivo', methods=['POST'])
@login_required
def excluir_arquivo_relatorio(id):
    """Exclui um arquivo individual do relatório (imagem_produto, ppt, ficha_tecnica)"""
    relatorio = Relatorio.query.get_or_404(id)
    campo = request.form.get('campo')

    campos_permitidos = ['imagem_produto', 'ppt_path', 'ficha_tecnica']
    if campo not in campos_permitidos:
        flash("Campo inválido.", "error")
        return redirect(url_for('editar_relatorio', id=id))

    try:
        file_path = getattr(relatorio, campo)
        if file_path:
            from utils import delete_file
            delete_file(file_path)
            setattr(relatorio, campo, None)
            db.session.commit()
            flash("Arquivo excluído com sucesso!", "success")
        else:
            flash("Nenhum arquivo para excluir.", "warning")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir o arquivo: {e}", "error")

    return redirect(url_for('editar_relatorio', id=id))


@app.route('/prova/<int:prova_id>/excluir_tabela', methods=['POST'])
@login_required
def excluir_tabela_medidas(prova_id):
    """Exclui a tabela de medidas de uma prova"""
    prova = Prova.query.get_or_404(prova_id)
    relatorio_id = prova.referencia.relatorio_id

    try:
        if prova.tabela_medidas_path:
            from utils import delete_file
            delete_file(prova.tabela_medidas_path)
            prova.tabela_medidas_path = None
            db.session.commit()
            flash("Tabela de medidas excluída com sucesso!", "success")
        else:
            flash("Nenhuma tabela de medidas para excluir.", "warning")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir a tabela de medidas: {e}", "error")

    redirect_to = request.form.get('redirect_to', '')
    if redirect_to == 'editar':
        return redirect(url_for('editar_relatorio', id=relatorio_id))
    return redirect(url_for('detalhes_relatorio', id=relatorio_id))


@app.route('/foto/<int:foto_id>/excluir', methods=['POST'])
@login_required
def excluir_foto(foto_id):
    """Exclui uma foto individual de uma prova"""
    foto = Foto.query.get_or_404(foto_id)
    relatorio_id = foto.prova.referencia.relatorio_id

    try:
        file_path = foto.file_path
        db.session.delete(foto)
        db.session.commit()

        if file_path:
            from utils import delete_file
            delete_file(file_path)

        flash("Foto excluída com sucesso!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir a foto: {e}", "error")

    redirect_to = request.form.get('redirect_to', '')
    if redirect_to == 'editar':
        return redirect(url_for('editar_relatorio', id=relatorio_id))
    return redirect(url_for('detalhes_relatorio', id=relatorio_id))


@app.route('/referencia/<int:referencia_id>/nova_prova', methods=['GET', 'POST'])
@login_required
def adicionar_nova_prova(referencia_id):
    referencia = Referencia.query.get_or_404(referencia_id)

    if request.method == 'POST':
        try:
            novo_numero_prova = request.form.get('numero_prova')
            tipo = referencia.tipo_categoria

            tabela_medidas_filename = save_file(request.files.get(f'tabela_medidas_{tipo}'))

            nova_prova = Prova(
                referencia_id=referencia.id,
                numero_prova=novo_numero_prova,
                tabela_medidas_path=tabela_medidas_filename,
                data_recebimento=request.form.get(f'data_recebimento_{tipo}'),
                tamanhos_recebidos=", ".join(request.form.getlist(f'tamanhos_recebidos_{tipo}')),
                info_medidas=request.form.get(f'info_medidas_{tipo}'),
                data_prova=request.form.get(f'data_prova_{tipo}'),
                time_qualidade=request.form.get(f'time_qualidade_{tipo}'),
                checklist_qualidade=", ".join(request.form.getlist(f'checklist_qualidade_{tipo}')),
                comentarios_qualidade=request.form.get(f'comentarios_qualidade_{tipo}'),
                obs_qualidade=request.form.get(f'obs_qualidade_{tipo}'),
                time_estilo=request.form.get(f'time_estilo_{tipo}'),
                checklist_estilo=", ".join(request.form.getlist(f'checklist_estilo_{tipo}')),
                comentarios_estilo=request.form.get(f'comentarios_estilo_{tipo}'),
                obs_estilo=request.form.get(f'obs_estilo_{tipo}'),
                time_modelagem=request.form.get(f'time_modelagem_{tipo}'),
                checklist_modelagem=", ".join(request.form.getlist(f'checklist_modelagem_{tipo}')),
                comentarios_modelagem=request.form.get(f'comentarios_modelagem_{tipo}'),
                obs_modelagem=request.form.get(f'obs_modelagem_{tipo}'),
                data_lacre=request.form.get(f'data_lacre_{tipo}'),
                numero_lacre=request.form.get(f'numero_lacre_{tipo}'),
                info_adicionais=request.form.get(f'info_adicionais_{tipo}')
            )
            db.session.add(nova_prova)
            db.session.flush()

            # Persistir respostas do checklist dinâmico para a nova prova
            # (form usa o tipo como sufixo em vez do prova_id, pois prova é nova)
            for cat in ('qualidade', 'estilo', 'modelagem'):
                _salvar_checklist_dinamico_para_prova(
                    nova_prova.id, cat,
                    request.form.getlist(f'checklist_dinamico_{cat}_{tipo}'),
                )

            campos_fotos = ['desenho', 'qualidade', 'estilo', 'modelagem']
            for contexto in campos_fotos:
                for file in request.files.getlist(f'fotos_{contexto}_{tipo}'):
                    filename = save_file(file)
                    if filename:
                        foto = Foto(prova_id=nova_prova.id, contexto=contexto, file_path=filename)
                        db.session.add(foto)

            # Upload direto de fotos amostra/prova_modelo (sem tamanho)
            for contexto in ['amostra', 'prova_modelo']:
                for file in request.files.getlist(f'fotos_{contexto}_{tipo}'):
                    filename = save_file(file)
                    if filename:
                        foto = Foto(prova_id=nova_prova.id, contexto=contexto, file_path=filename)
                        db.session.add(foto)

            # Upload por tamanho (opcional)
            for tamanho in request.form.getlist(f'tamanhos_recebidos_{tipo}'):
                for contexto in ['amostra', 'prova_modelo']:
                    for file in request.files.getlist(f'fotos_{contexto}_{tipo}_{tamanho.replace(" ", "")}'):
                        filename = save_file(file)
                        if filename:
                            foto = Foto(prova_id=nova_prova.id, contexto=contexto, tamanho=tamanho, file_path=filename)
                            db.session.add(foto)

            db.session.commit()
            flash(f"{novo_numero_prova}ª prova adicionada com sucesso!", "success")

            gerar_e_salvar_pdf(referencia.relatorio_id, evento="ATUALIZADO")

            return redirect(url_for('detalhes_relatorio', id=referencia.relatorio_id))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao salvar a nova prova: {e}", "error")
            return redirect(url_for('detalhes_relatorio', id=referencia.relatorio_id))
    
    ultima_prova = Prova.query.filter_by(referencia_id=referencia_id).order_by(desc(Prova.numero_prova)).first()
    novo_numero_prova = ultima_prova.numero_prova + 1 if ultima_prova else 1

    return render_template('nova_prova.html', referencia=referencia, novo_numero_prova=novo_numero_prova)


@app.route('/exportar/excel')
@login_required
def exportar_relatorios_excel():
    """Exporta todos os relatórios para Excel"""
    try:
        relatorios = Relatorio.query.order_by(desc(Relatorio.created_at)).all()
        relatorios_data = []

        for relatorio in relatorios:
            # Contar referências
            num_referencias = Referencia.query.filter_by(relatorio_id=relatorio.id).count()

            # Obter status atual
            ultima_prova = Prova.query.join(Referencia).filter(
                Referencia.relatorio_id == relatorio.id
            ).order_by(desc(Prova.numero_prova)).first()

            relatorios_data.append({
                'id': relatorio.id,
                'colecao': relatorio.colecao or '',
                'descricao_geral': relatorio.descricao_geral or '',
                'num_referencias': num_referencias,
                'status_geral': ultima_prova.status if ultima_prova else 'Novo',
                'data_criacao': relatorio.created_at.strftime('%d/%m/%Y %H:%M') if relatorio.created_at else ''
            })

        filename = export_relatorios_to_excel(relatorios_data)

        flash(f'Relatórios exportados com sucesso!', 'success')
        return send_from_directory(app.config['PDF_FOLDER'], filename, as_attachment=True)

    except Exception as e:
        flash(f'Erro ao exportar relatórios: {e}', 'error')
        return redirect(url_for('dashboard'))


@app.route('/relatorio/<int:id>/excel')
@login_required
def exportar_relatorio_excel(id):
    """Exporta detalhes de um relatório específico para Excel"""
    try:
        relatorio = Relatorio.query.get_or_404(id)

        # Preparar dados do relatório
        relatorio_data = {
            'id': relatorio.id,
            'colecao': relatorio.colecao,
            'descricao_geral': relatorio.descricao_geral
        }

        # Preparar referências com provas
        referencias = []
        for ref in relatorio.referencias:
            ref_data = {
                'numero_ref': ref.numero_ref,
                'tipo': ref.tipo_categoria,
                'origem': ref.origem or '',
                'fornecedor': (ref.fornecedor_obj.nome if ref.fornecedor_obj else (ref.fornecedor or '')),
                'fornecedor_pais': (ref.fornecedor_obj.pais if ref.fornecedor_obj and ref.fornecedor_obj.pais else ''),
                'fornecedor_contato': (ref.fornecedor_obj.contato if ref.fornecedor_obj else (ref.fornecedor_contato or '')),
                'materia_prima': ref.materia_prima or '',
                'composicao': ref.composicao or '',
                'gramatura': ref.gramatura or '',
                'aviamentos': ref.aviamentos or '',
                'observacoes': ref.observacoes or '',
                'provas': []
            }

            for prova in ref.provas:
                ref_data['provas'].append({
                    'numero_prova': prova.numero_prova,
                    'status': prova.status,
                    'data_recebimento': prova.data_recebimento,
                    'data_prova': prova.data_prova,
                    'tamanhos_recebidos': prova.tamanhos_recebidos
                })

            referencias.append(ref_data)

        filename = export_detalhes_to_excel(relatorio_data, referencias)

        flash(f'Relatório exportado para Excel!', 'success')
        return send_from_directory(app.config['PDF_FOLDER'], filename, as_attachment=True)

    except Exception as e:
        flash(f'Erro ao exportar relatório: {e}', 'error')
        return redirect(url_for('detalhes_relatorio', id=id))


@app.route('/importar/excel/modelo')
@login_required
def download_modelo_excel():
    """Gera e retorna um arquivo Excel modelo para importação (formato Analytics)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    import tempfile

    wb = Workbook()

    # Estilos compartilhados
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="e6007e", end_color="e6007e", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========================================
    # ABA 1: INFORMAÇÕES GERAIS (Relatórios)
    # ========================================
    ws1 = wb.active
    ws1.title = "Informações Gerais"

    headers1 = ['Descrição', 'Linha', 'Coleção', 'Temporada', 'Ano', 'Status Geral']
    for col_idx, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Linha de exemplo
    exemplo1 = ['VESTIDO FLORAL', 'PRAIA', 'VERÃO 2025', 'VERÃO', 2025, 'EM ANDAMENTO']
    for col_idx, valor in enumerate(exemplo1, 1):
        cell = ws1.cell(row=2, column=col_idx, value=valor)
        cell.alignment = cell_alignment
        cell.border = thin_border

    col_widths1 = [30, 15, 20, 15, 10, 18]
    for col_idx, width in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # Data validation — Linha (coluna B)
    dv_linha = DataValidation(
        type="list",
        formula1='"PRAIA,ACESSÓRIO,LINGERIE,MEIAS,HOMEWEAR"',
        allow_blank=True
    )
    dv_linha.prompt = "Selecione a linha do produto"
    dv_linha.promptTitle = "Linha"
    dv_linha.showInputMessage = True
    ws1.add_data_validation(dv_linha)
    dv_linha.add('B2:B1048576')

    # Data validation — Status Geral (coluna F)
    dv_status1 = DataValidation(
        type="list",
        formula1='"EM ANDAMENTO,APROVADA,REPROVADA,COMITÊ"',
        allow_blank=True
    )
    dv_status1.prompt = "Selecione o status"
    dv_status1.promptTitle = "Status Geral"
    dv_status1.showInputMessage = True
    ws1.add_data_validation(dv_status1)
    dv_status1.add('F2:F1048576')

    # ========================================
    # ABA 2: DADOS DETALHADOS (Referências + Provas)
    # ========================================
    ws2 = wb.create_sheet("Dados Detalhados")

    headers2 = [
        "Coleção", "Descrição", "Referência", "Categoria",
        "Fornecedor", "Nº Prova", "Status", "Data Prova",
        "Data Recebimento", "Tamanhos", "Time Qualidade",
        "Time Estilo", "Time Modelagem"
    ]
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Linha de exemplo
    exemplo2 = [
        'VERÃO 2025', 'VESTIDO FLORAL', 'REF-001', 'ADULTO',
        'FORNECEDOR X', 1, 'EM ANDAMENTO', '15/03/2025',
        '10/03/2025', 'P, M, G', 'ANA SILVA',
        'CARLOS LIMA', 'MARIA SOUZA'
    ]
    for col_idx, valor in enumerate(exemplo2, 1):
        cell = ws2.cell(row=2, column=col_idx, value=valor)
        cell.alignment = cell_alignment
        cell.border = thin_border

    col_widths2 = [20, 30, 15, 12, 20, 10, 16, 14, 16, 15, 18, 18, 18]
    for col_idx, width in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    # Data validation — Categoria (coluna D)
    dv_categoria = DataValidation(
        type="list",
        formula1='"BABY,KIDS,TEEN,ADULTO"',
        allow_blank=True
    )
    dv_categoria.prompt = "Selecione a categoria"
    dv_categoria.promptTitle = "Categoria"
    dv_categoria.showInputMessage = True
    ws2.add_data_validation(dv_categoria)
    dv_categoria.add('D2:D1048576')

    # Data validation — Status (coluna G)
    dv_status2 = DataValidation(
        type="list",
        formula1='"EM ANDAMENTO,APROVADA,REPROVADA,COMITÊ"',
        allow_blank=True
    )
    dv_status2.prompt = "Selecione o status da prova"
    dv_status2.promptTitle = "Status"
    dv_status2.showInputMessage = True
    ws2.add_data_validation(dv_status2)
    dv_status2.add('G2:G1048576')

    # Salvar e retornar
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    return send_file(
        tmp_path,
        as_attachment=True,
        download_name='modelo_importacao.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def _formatar_data_excel(valor):
    """Converte datetime do Excel para string dd/mm/yyyy."""
    if valor is None:
        return None
    if hasattr(valor, 'strftime'):
        return valor.strftime('%d/%m/%Y')
    return str(valor).strip() or None


def _to_int_or_none(v):
    if v is None or v == '':
        return None
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return None


def _parsear_planilha_importacao(filepath):
    """Lê o XLSX e retorna (parsed, erros).
    parsed = {'gerais': [...], 'detalhados': [...]} com cada linha como dict.
    Inclui IDs quando presentes para detectar UPSERT vs INSERT.
    """
    from openpyxl import load_workbook
    wb = load_workbook(filepath, data_only=True)
    parsed = {'gerais': [], 'detalhados': []}
    erros = []

    # ABA Informações Gerais
    if "Informações Gerais" in wb.sheetnames:
        ws1 = wb["Informações Gerais"]
    else:
        ws1 = wb.active

    headers1 = [c.value for c in ws1[1]]
    for row_idx, row in enumerate(ws1.iter_rows(min_row=2, values_only=True), start=2):
        try:
            dados = dict(zip(headers1, row))
            if not dados.get('Descrição'):
                continue
            parsed['gerais'].append({
                'linha_planilha': row_idx,
                'id_relatorio': _to_int_or_none(dados.get('ID Relatorio')),
                'descricao': str(dados.get('Descrição', '')).strip().upper(),
                'linha': str(dados.get('Linha', '')).strip().upper() if dados.get('Linha') else None,
                'colecao': str(dados.get('Coleção', '')).strip().upper() if dados.get('Coleção') else None,
                'temporada': str(dados.get('Temporada', '')).strip().upper() if dados.get('Temporada') else None,
                'ano': _to_int_or_none(dados.get('Ano')),
                'status_geral': str(dados.get('Status Geral', 'EM ANDAMENTO')).strip().upper(),
            })
        except Exception as e:
            erros.append({'sheet': 'Informações Gerais', 'linha': row_idx, 'mensagem': str(e)})

    # ABA Dados Detalhados
    if "Dados Detalhados" in wb.sheetnames:
        ws2 = wb["Dados Detalhados"]
        headers2 = [c.value for c in ws2[1]]
        for row_idx, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
            try:
                dados = dict(zip(headers2, row))
                if not dados.get('Referência') and not dados.get('Descrição'):
                    continue
                parsed['detalhados'].append({
                    'linha_planilha': row_idx,
                    'id_relatorio': _to_int_or_none(dados.get('ID Relatorio')),
                    'id_referencia': _to_int_or_none(dados.get('ID Referencia')),
                    'id_prova': _to_int_or_none(dados.get('ID Prova')),
                    'colecao': str(dados.get('Coleção', '')).strip().upper() if dados.get('Coleção') else '',
                    'descricao': str(dados.get('Descrição', '')).strip().upper() if dados.get('Descrição') else '',
                    'numero_ref': str(dados.get('Referência', '')).strip().upper() if dados.get('Referência') else '',
                    'categoria': str(dados.get('Categoria', '')).strip().upper() if dados.get('Categoria') else 'ADULTO',
                    'fornecedor': str(dados.get('Fornecedor', '')).strip().upper() if dados.get('Fornecedor') else None,
                    'numero_prova': _to_int_or_none(dados.get('Nº Prova')) or 1,
                    'status': str(dados.get('Status', 'EM ANDAMENTO')).strip().upper(),
                    'data_prova': _formatar_data_excel(dados.get('Data Prova')),
                    'data_recebimento': _formatar_data_excel(dados.get('Data Recebimento')),
                    'tamanhos_recebidos': str(dados.get('Tamanhos', '')).strip() if dados.get('Tamanhos') else None,
                    'time_qualidade': str(dados.get('Time Qualidade', '')).strip() if dados.get('Time Qualidade') else None,
                    'time_estilo': str(dados.get('Time Estilo', '')).strip() if dados.get('Time Estilo') else None,
                    'time_modelagem': str(dados.get('Time Modelagem', '')).strip() if dados.get('Time Modelagem') else None,
                })
            except Exception as e:
                erros.append({'sheet': 'Dados Detalhados', 'linha': row_idx, 'mensagem': str(e)})

    return parsed, erros


def _calcular_summary(parsed):
    """Pré-calcula quantos registros serão criados vs atualizados."""
    summary = {
        'criar_relatorios': 0, 'atualizar_relatorios': 0,
        'criar_referencias': 0, 'atualizar_referencias': 0,
        'criar_provas': 0, 'atualizar_provas': 0,
    }
    for linha in parsed['gerais']:
        if linha['id_relatorio'] and Relatorio.query.get(linha['id_relatorio']):
            summary['atualizar_relatorios'] += 1
        else:
            summary['criar_relatorios'] += 1
    for linha in parsed['detalhados']:
        if linha.get('id_referencia') and Referencia.query.get(linha['id_referencia']):
            summary['atualizar_referencias'] += 1
        elif linha.get('numero_ref'):
            summary['criar_referencias'] += 1
        if linha.get('id_prova') and Prova.query.get(linha['id_prova']):
            summary['atualizar_provas'] += 1
        else:
            summary['criar_provas'] += 1
    return summary


@app.route('/importar/excel', methods=['GET', 'POST'])
@login_required
def importar_relatorios_excel():
    """Etapa 1: recebe upload, parseia, valida, cria ImportJob e redireciona para revisar."""
    if request.method != 'POST':
        return redirect(url_for('dashboard'))

    arquivo = request.files.get('arquivo_excel')
    if not arquivo or arquivo.filename == '':
        flash('Nenhum arquivo selecionado!', 'error')
        return redirect(url_for('dashboard'))

    if not arquivo.filename.endswith(('.xlsx', '.xls')):
        flash('Formato inválido! Use apenas arquivos Excel (.xlsx ou .xls)', 'error')
        return redirect(url_for('dashboard'))

    import tempfile as _tempfile
    with _tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        arquivo.save(temp_file.name)
        temp_path = temp_file.name

    try:
        parsed, erros = _parsear_planilha_importacao(temp_path)
        summary = _calcular_summary(parsed)
    except Exception as e:
        try:
            os.unlink(temp_path)
        except OSError:
            pass
        flash(f'Erro ao ler arquivo: {e}', 'error')
        return redirect(url_for('dashboard'))

    job = ImportJob(
        user_id=current_user.id if current_user.is_authenticated else None,
        arquivo_original=arquivo.filename,
        arquivo_temp_path=temp_path,
        status='validated',
    )
    job.set_parsed(parsed)
    job.set_erros(erros)
    job.set_summary(summary)
    db.session.add(job)
    db.session.commit()

    return redirect(url_for('importar_excel_revisar', job_id=job.id))


@app.route('/importar/excel/<int:job_id>/revisar')
@login_required
def importar_excel_revisar(job_id):
    """Etapa 2: exibe revisão dos dados parseados antes de commitar."""
    job = ImportJob.query.get_or_404(job_id)
    if job.user_id and job.user_id != current_user.id and not current_user.is_admin:
        abort(403)
    return render_template(
        'importar_excel_revisar.html',
        job=job,
        parsed=job.get_parsed(),
        erros=job.get_erros(),
        summary=job.get_summary(),
    )


@app.route('/importar/excel/<int:job_id>/confirmar', methods=['POST'])
@login_required
def importar_excel_confirmar(job_id):
    """Etapa 3: aplica criações/atualizações no banco."""
    job = ImportJob.query.get_or_404(job_id)
    if job.user_id and job.user_id != current_user.id and not current_user.is_admin:
        abort(403)
    if job.status != 'validated':
        flash('Este job já foi processado ou cancelado.', 'warning')
        return redirect(url_for('dashboard'))

    parsed = job.get_parsed()
    erros_runtime = []

    criados = {'relatorios': 0, 'referencias': 0, 'provas': 0}
    atualizados = {'relatorios': 0, 'referencias': 0, 'provas': 0}

    try:
        # ETAPA A: Informações Gerais (Relatórios)
        relatorios_map = {}  # chave descricao|colecao -> Relatorio
        for linha in parsed.get('gerais', []):
            try:
                rel = None
                if linha.get('id_relatorio'):
                    rel = Relatorio.query.get(linha['id_relatorio'])
                if rel:
                    rel.descricao_geral = linha['descricao']
                    rel.linha = linha.get('linha')
                    rel.colecao = linha.get('colecao')
                    rel.temporada = linha.get('temporada')
                    rel.ano = linha.get('ano')
                    rel.status_geral = linha.get('status_geral') or rel.status_geral
                    atualizados['relatorios'] += 1
                else:
                    rel = Relatorio(
                        codigo=gerar_codigo_relatorio(),
                        descricao_geral=linha['descricao'],
                        linha=linha.get('linha'),
                        colecao=linha.get('colecao'),
                        temporada=linha.get('temporada'),
                        ano=linha.get('ano'),
                        status_geral=linha.get('status_geral') or 'EM ANDAMENTO',
                        created_by=current_user.id if current_user.is_authenticated else None,
                    )
                    db.session.add(rel)
                    db.session.flush()
                    criados['relatorios'] += 1
                chave = f"{linha['descricao']}|{linha.get('colecao') or ''}"
                relatorios_map[chave] = rel
            except Exception as e:
                erros_runtime.append({'sheet': 'Informações Gerais', 'linha': linha.get('linha_planilha'), 'mensagem': str(e)})

        # ETAPA B: Dados Detalhados (Referências + Provas)
        refs_map = {}
        for linha in parsed.get('detalhados', []):
            try:
                # Achar relatório
                rel = None
                if linha.get('id_relatorio'):
                    rel = Relatorio.query.get(linha['id_relatorio'])
                if not rel:
                    chave = f"{linha.get('descricao', '')}|{linha.get('colecao', '')}"
                    rel = relatorios_map.get(chave)
                if not rel and linha.get('descricao'):
                    rel = Relatorio.query.filter(
                        db.func.upper(Relatorio.descricao_geral) == linha['descricao'],
                        db.func.upper(db.func.coalesce(Relatorio.colecao, '')) == linha.get('colecao', ''),
                    ).first()
                if not rel:
                    erros_runtime.append({
                        'sheet': 'Dados Detalhados',
                        'linha': linha.get('linha_planilha'),
                        'mensagem': f"Relatório não encontrado para '{linha.get('descricao')}' / '{linha.get('colecao')}'",
                    })
                    continue

                # Referência (upsert)
                ref = None
                if linha.get('id_referencia'):
                    ref = Referencia.query.get(linha['id_referencia'])
                if ref:
                    if linha.get('numero_ref'):
                        ref.numero_ref = linha['numero_ref']
                    if linha.get('categoria'):
                        ref.tipo_categoria = linha['categoria'].lower() if linha['categoria'].upper() in ('BABY', 'KIDS', 'TEEN', 'ADULTO') else ref.tipo_categoria
                    if linha.get('fornecedor'):
                        ref.fornecedor = linha['fornecedor']
                    atualizados['referencias'] += 1
                else:
                    chave_ref = f"{rel.id}|{linha.get('numero_ref', '')}"
                    if chave_ref in refs_map:
                        ref = refs_map[chave_ref]
                    else:
                        ref = Referencia(
                            relatorio_id=rel.id,
                            numero_ref=linha.get('numero_ref'),
                            tipo_categoria=(linha.get('categoria') or 'ADULTO').lower(),
                            fornecedor=linha.get('fornecedor'),
                        )
                        db.session.add(ref)
                        db.session.flush()
                        refs_map[chave_ref] = ref
                        criados['referencias'] += 1

                # Prova (upsert)
                prova = None
                if linha.get('id_prova'):
                    prova = Prova.query.get(linha['id_prova'])
                if prova:
                    prova.numero_prova = linha.get('numero_prova') or prova.numero_prova
                    prova.status = linha.get('status') or prova.status
                    prova.data_prova = linha.get('data_prova')
                    prova.data_recebimento = linha.get('data_recebimento')
                    if linha.get('tamanhos_recebidos'):
                        prova.tamanhos_recebidos = linha['tamanhos_recebidos']
                    if linha.get('time_qualidade'):
                        prova.time_qualidade = linha['time_qualidade']
                    if linha.get('time_estilo'):
                        prova.time_estilo = linha['time_estilo']
                    if linha.get('time_modelagem'):
                        prova.time_modelagem = linha['time_modelagem']
                    atualizados['provas'] += 1
                else:
                    prova = Prova(
                        referencia_id=ref.id,
                        numero_prova=linha.get('numero_prova') or 1,
                        status=linha.get('status') or 'EM ANDAMENTO',
                        data_prova=linha.get('data_prova'),
                        data_recebimento=linha.get('data_recebimento'),
                        tamanhos_recebidos=linha.get('tamanhos_recebidos'),
                        time_qualidade=linha.get('time_qualidade'),
                        time_estilo=linha.get('time_estilo'),
                        time_modelagem=linha.get('time_modelagem'),
                    )
                    db.session.add(prova)
                    criados['provas'] += 1
            except Exception as e:
                erros_runtime.append({'sheet': 'Dados Detalhados', 'linha': linha.get('linha_planilha'), 'mensagem': str(e)})

        db.session.commit()
        from datetime import datetime as _dt_now
        erros_parse_originais = job.get_erros()
        todos_erros = erros_runtime + erros_parse_originais
        job.status = 'committed'
        job.confirmed_at = _dt_now.utcnow()
        if erros_runtime:
            job.set_erros(todos_erros)
        db.session.commit()

        registrar_log(
            acao='importar',
            entidade_tipo='import_job',
            entidade_id=job.id,
            entidade_descricao=f'Importação Excel #{job.id}',
            detalhes=f"criados={criados} atualizados={atualizados} erros={len(erros_runtime)}",
        )

        # Limpar arquivo temporário
        if job.arquivo_temp_path:
            try:
                os.unlink(job.arquivo_temp_path)
            except OSError:
                pass

        return render_template(
            'importar_excel_resultado.html',
            job=job,
            criados=criados,
            atualizados=atualizados,
            erros=todos_erros,
        )
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao confirmar importação: {e}', 'error')
        return redirect(url_for('importar_excel_revisar', job_id=job.id))


@app.route('/importar/excel/<int:job_id>/cancelar', methods=['POST'])
@login_required
def importar_excel_cancelar(job_id):
    """Cancela um job de importação e limpa arquivo temporário."""
    job = ImportJob.query.get_or_404(job_id)
    if job.user_id and job.user_id != current_user.id and not current_user.is_admin:
        abort(403)
    if job.arquivo_temp_path:
        try:
            os.unlink(job.arquivo_temp_path)
        except OSError:
            pass
    job.status = 'cancelled'
    db.session.commit()
    flash('Importação cancelada.', 'info')
    return redirect(url_for('dashboard'))


@app.route('/exportar/excel/editavel')
@login_required
def exportar_excel_editavel():
    """Exporta todos os relatórios em formato editável (com IDs ocultos)
    para o fluxo round-trip (export → editar → reimportar)."""
    relatorios = Relatorio.query.filter_by(is_active=True).all()
    try:
        filepath = export_editavel(relatorios)
    except Exception as e:
        flash(f'Erro ao exportar: {e}', 'error')
        return redirect(url_for('dashboard'))

    from datetime import datetime as _dt_now
    timestamp = _dt_now.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        filepath,
        as_attachment=True,
        download_name=f'relatorios_editavel_{timestamp}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ========================================
# PÁGINA DE RELATÓRIOS E ANALYTICS
# ========================================

@app.route('/analytics')
@login_required
def analytics():
    """Página de relatórios e analytics com filtros"""
    from datetime import datetime, timedelta

    # Obter parâmetros de filtro
    filtro_status = request.args.get('status', '')
    filtro_categoria = request.args.get('categoria', '')
    filtro_colecao = request.args.get('colecao', '')
    filtro_fornecedor = request.args.get('fornecedor', '')
    filtro_referencia = request.args.get('referencia', '')
    filtro_data_inicio = request.args.get('data_inicio', '')
    filtro_data_fim = request.args.get('data_fim', '')

    # Query base de provas
    query_provas = Prova.query.join(Referencia).join(Relatorio)

    # Aplicar filtros
    if filtro_status:
        query_provas = query_provas.filter(Prova.status == filtro_status)
    if filtro_categoria:
        query_provas = query_provas.filter(Referencia.tipo_categoria == filtro_categoria)
    if filtro_colecao:
        query_provas = query_provas.filter(Relatorio.colecao == filtro_colecao)
    if filtro_fornecedor:
        query_provas = query_provas.filter(Referencia.fornecedor == filtro_fornecedor)
    if filtro_referencia:
        # Busca parcial por número de referência (case-insensitive)
        query_provas = query_provas.filter(Referencia.numero_ref.ilike(f'%{filtro_referencia}%'))

    # Filtro de data
    if filtro_data_inicio:
        try:
            data_inicio = datetime.strptime(filtro_data_inicio, '%Y-%m-%d')
            query_provas = query_provas.filter(Prova.created_at >= data_inicio)
        except:
            pass
    if filtro_data_fim:
        try:
            data_fim = datetime.strptime(filtro_data_fim, '%Y-%m-%d')
            data_fim = data_fim + timedelta(days=1)  # Incluir o dia inteiro
            query_provas = query_provas.filter(Prova.created_at < data_fim)
        except:
            pass

    # ========================================
    # ESTATÍSTICAS GERAIS (sem filtros)
    # ========================================
    total_relatorios = Relatorio.query.count()
    total_referencias = Referencia.query.count()
    total_provas = Prova.query.count()

    # Por status (sem filtros) - Valores em MAIÚSCULAS após padronização
    provas_aprovadas = Prova.query.filter_by(status='APROVADA').count()
    provas_reprovadas = Prova.query.filter_by(status='REPROVADA').count()
    provas_em_andamento = Prova.query.filter_by(status='EM ANDAMENTO').count()
    provas_comite = Prova.query.filter_by(status='COMITÊ').count()

    # Taxa de aprovação
    taxa_aprovacao = round((provas_aprovadas / total_provas * 100) if total_provas > 0 else 0, 1)
    taxa_reprovacao = round((provas_reprovadas / total_provas * 100) if total_provas > 0 else 0, 1)

    # Provas com retrabalho
    provas_retrabalho = Prova.query.filter(Prova.numero_prova > 1).count()
    taxa_retrabalho = round((provas_retrabalho / total_provas * 100) if total_provas > 0 else 0, 1)

    # ========================================
    # ESTATÍSTICAS FILTRADAS
    # ========================================
    provas_filtradas = query_provas.all()
    total_filtrado = len(provas_filtradas)

    filtrado_aprovadas = sum(1 for p in provas_filtradas if p.status == 'APROVADA')
    filtrado_reprovadas = sum(1 for p in provas_filtradas if p.status == 'REPROVADA')
    filtrado_em_andamento = sum(1 for p in provas_filtradas if p.status == 'EM ANDAMENTO')
    filtrado_comite = sum(1 for p in provas_filtradas if p.status == 'COMITÊ')

    taxa_aprovacao_filtrada = round((filtrado_aprovadas / total_filtrado * 100) if total_filtrado > 0 else 0, 1)

    # ========================================
    # DADOS PARA GRÁFICOS
    # ========================================

    # Por categoria
    categorias_stats = db.session.query(
        Referencia.tipo_categoria,
        db.func.count(Referencia.id)
    ).group_by(Referencia.tipo_categoria).all()

    # Por status
    status_stats = db.session.query(
        Prova.status,
        db.func.count(Prova.id)
    ).group_by(Prova.status).all()

    # Por fornecedor (top 10)
    fornecedores_stats = db.session.query(
        Referencia.fornecedor,
        db.func.count(Referencia.id)
    ).filter(Referencia.fornecedor.isnot(None), Referencia.fornecedor != '').group_by(
        Referencia.fornecedor
    ).order_by(db.func.count(Referencia.id).desc()).limit(10).all()

    # Por coleção
    colecoes_stats = db.session.query(
        Relatorio.colecao,
        db.func.count(Relatorio.id)
    ).filter(Relatorio.colecao.isnot(None), Relatorio.colecao != '').group_by(
        Relatorio.colecao
    ).order_by(db.func.count(Relatorio.id).desc()).all()

    # Relatórios por mês (últimos 12 meses)
    doze_meses_atras = datetime.utcnow() - timedelta(days=365)
    relatorios_por_mes = db.session.query(
        db.func.strftime('%Y-%m', Relatorio.created_at).label('mes'),
        db.func.count(Relatorio.id)
    ).filter(Relatorio.created_at >= doze_meses_atras).group_by('mes').order_by('mes').all()

    # ========================================
    # INSIGHTS INTELIGENTES
    # ========================================
    insights = []

    # Insight de aprovação
    if taxa_aprovacao >= 80:
        insights.append({
            'tipo': 'success',
            'icone': 'bi-trophy-fill',
            'titulo': 'Excelente Taxa de Aprovação',
            'mensagem': f'{taxa_aprovacao}% das provas foram aprovadas. Parabéns pela qualidade!'
        })
    elif taxa_aprovacao >= 60:
        insights.append({
            'tipo': 'warning',
            'icone': 'bi-exclamation-triangle-fill',
            'titulo': 'Taxa de Aprovação Moderada',
            'mensagem': f'{taxa_aprovacao}% de aprovação. Há espaço para melhorias no processo.'
        })
    else:
        insights.append({
            'tipo': 'danger',
            'icone': 'bi-x-circle-fill',
            'titulo': 'Atenção: Baixa Aprovação',
            'mensagem': f'Apenas {taxa_aprovacao}% aprovadas. Recomenda-se revisar fornecedores e processos.'
        })

    # Insight de retrabalho
    if taxa_retrabalho > 30:
        insights.append({
            'tipo': 'warning',
            'icone': 'bi-arrow-repeat',
            'titulo': 'Alto Índice de Retrabalho',
            'mensagem': f'{taxa_retrabalho}% das provas precisaram de mais de uma tentativa.'
        })
    elif taxa_retrabalho > 0:
        insights.append({
            'tipo': 'info',
            'icone': 'bi-arrow-repeat',
            'titulo': 'Retrabalho Controlado',
            'mensagem': f'{taxa_retrabalho}% das provas necessitaram retrabalho.'
        })

    # Insight de provas pendentes
    if provas_em_andamento > 10:
        insights.append({
            'tipo': 'info',
            'icone': 'bi-hourglass-split',
            'titulo': 'Provas Aguardando Análise',
            'mensagem': f'{provas_em_andamento} provas em andamento aguardando conclusão.'
        })

    # Insight de comitê
    if provas_comite > 0:
        insights.append({
            'tipo': 'primary',
            'icone': 'bi-people-fill',
            'titulo': 'Provas para Comitê',
            'mensagem': f'{provas_comite} provas aguardando decisão do comitê.'
        })

    # Top fornecedor
    if fornecedores_stats:
        top_fornecedor = fornecedores_stats[0]
        insights.append({
            'tipo': 'secondary',
            'icone': 'bi-building',
            'titulo': 'Fornecedor mais Ativo',
            'mensagem': f'"{top_fornecedor[0]}" com {top_fornecedor[1]} referências.'
        })

    # ========================================
    # OPÇÕES PARA FILTROS (dropdowns)
    # ========================================
    opcoes_status = ['EM ANDAMENTO', 'APROVADA', 'REPROVADA', 'COMITÊ']
    opcoes_categorias = ['BABY', 'KIDS', 'TEEN', 'ADULTO']

    # Coleções únicas
    opcoes_colecoes = [c[0] for c in db.session.query(Relatorio.colecao).filter(
        Relatorio.colecao.isnot(None), Relatorio.colecao != ''
    ).distinct().order_by(Relatorio.colecao).all()]

    # Fornecedores únicos
    opcoes_fornecedores = [f[0] for f in db.session.query(Referencia.fornecedor).filter(
        Referencia.fornecedor.isnot(None), Referencia.fornecedor != ''
    ).distinct().order_by(Referencia.fornecedor).all()]

    # ========================================
    # TABELA DE DADOS FILTRADOS
    # ========================================
    dados_tabela = []
    for prova in provas_filtradas:
        ref = prova.referencia
        rel = ref.relatorio
        dados_tabela.append({
            'relatorio_id': rel.id,
            'colecao': rel.colecao or '-',
            'descricao': rel.descricao_geral or '-',
            'referencia': ref.numero_ref or '-',
            'categoria': ref.tipo_categoria,
            'fornecedor': ref.fornecedor or '-',
            'numero_prova': prova.numero_prova,
            'status': prova.status,
            'data_prova': prova.data_prova or '-'
        })

    return render_template('analytics.html',
        # Estatísticas gerais
        total_relatorios=total_relatorios,
        total_referencias=total_referencias,
        total_provas=total_provas,
        provas_aprovadas=provas_aprovadas,
        provas_reprovadas=provas_reprovadas,
        provas_em_andamento=provas_em_andamento,
        provas_comite=provas_comite,
        taxa_aprovacao=taxa_aprovacao,
        taxa_reprovacao=taxa_reprovacao,
        taxa_retrabalho=taxa_retrabalho,

        # Estatísticas filtradas
        total_filtrado=total_filtrado,
        filtrado_aprovadas=filtrado_aprovadas,
        filtrado_reprovadas=filtrado_reprovadas,
        filtrado_em_andamento=filtrado_em_andamento,
        filtrado_comite=filtrado_comite,
        taxa_aprovacao_filtrada=taxa_aprovacao_filtrada,

        # Dados para gráficos
        categorias_stats=dict(categorias_stats),
        status_stats=dict(status_stats),
        fornecedores_stats=fornecedores_stats,
        colecoes_stats=colecoes_stats,
        relatorios_por_mes=relatorios_por_mes,

        # Insights
        insights=insights,

        # Opções para filtros
        opcoes_status=opcoes_status,
        opcoes_categorias=opcoes_categorias,
        opcoes_colecoes=opcoes_colecoes,
        opcoes_fornecedores=opcoes_fornecedores,

        # Filtros ativos
        filtro_status=filtro_status,
        filtro_categoria=filtro_categoria,
        filtro_colecao=filtro_colecao,
        filtro_fornecedor=filtro_fornecedor,
        filtro_referencia=filtro_referencia,
        filtro_data_inicio=filtro_data_inicio,
        filtro_data_fim=filtro_data_fim,

        # Dados da tabela
        dados_tabela=dados_tabela
    )


@app.route('/api/analytics/charts')
@login_required
def api_analytics_charts():
    """
    Endpoint API para fornecer dados dos gráficos
    Retorna JSON com todos os dados necessários para visualizações
    """
    from datetime import datetime, timedelta

    try:
        # Obter parâmetros de filtro (opcional)
        filtro_status = request.args.get('status', '')
        filtro_categoria = request.args.get('categoria', '')
        filtro_colecao = request.args.get('colecao', '')
        filtro_fornecedor = request.args.get('fornecedor', '')

        # Query base
        query_provas = Prova.query.join(Referencia).join(Relatorio)

        # Aplicar filtros se fornecidos
        if filtro_status:
            query_provas = query_provas.filter(Prova.status == filtro_status)
        if filtro_categoria:
            query_provas = query_provas.filter(Referencia.tipo_categoria == filtro_categoria)
        if filtro_colecao:
            query_provas = query_provas.filter(Relatorio.colecao == filtro_colecao)
        if filtro_fornecedor:
            query_provas = query_provas.filter(Referencia.fornecedor == filtro_fornecedor)

        # ========================================
        # 1. DISTRIBUIÇÃO POR STATUS (Pie/Doughnut)
        # ========================================
        status_stats = db.session.query(
            Prova.status,
            db.func.count(Prova.id)
        ).group_by(Prova.status).all()

        status_chart = {
            'labels': [status for status, _ in status_stats],
            'values': [count for _, count in status_stats]
        }

        # ========================================
        # 2. TOP 10 FORNECEDORES (Bar Horizontal)
        # ========================================
        fornecedores_stats = db.session.query(
            Referencia.fornecedor,
            db.func.count(Referencia.id)
        ).filter(
            Referencia.fornecedor.isnot(None),
            Referencia.fornecedor != ''
        ).group_by(
            Referencia.fornecedor
        ).order_by(
            db.func.count(Referencia.id).desc()
        ).limit(10).all()

        suppliers_chart = {
            'suppliers': [forn for forn, _ in fornecedores_stats],
            'counts': [count for _, count in fornecedores_stats]
        }

        # ========================================
        # 3. TIMELINE - RELATÓRIOS POR MÊS (Area Chart)
        # ========================================
        doze_meses_atras = datetime.utcnow() - timedelta(days=365)
        relatorios_por_mes = db.session.query(
            db.func.strftime('%Y-%m', Relatorio.created_at).label('mes'),
            db.func.count(Relatorio.id)
        ).filter(
            Relatorio.created_at >= doze_meses_atras
        ).group_by('mes').order_by('mes').all()

        # Formatar meses para display (Jan, Fev, Mar...)
        meses_map = {
            '01': 'Jan', '02': 'Fev', '03': 'Mar', '04': 'Abr',
            '05': 'Mai', '06': 'Jun', '07': 'Jul', '08': 'Ago',
            '09': 'Set', '10': 'Out', '11': 'Nov', '12': 'Dez'
        }

        timeline_chart = {
            'months': [meses_map.get(mes.split('-')[1], mes) if mes else '' for mes, _ in relatorios_por_mes],
            'counts': [count for _, count in relatorios_por_mes]
        }

        # ========================================
        # 4. DISTRIBUIÇÃO POR CATEGORIA (Bar Vertical)
        # ========================================
        categorias_stats = db.session.query(
            Referencia.tipo_categoria,
            db.func.count(Referencia.id)
        ).group_by(Referencia.tipo_categoria).all()

        category_chart = {
            'categories': [cat for cat, _ in categorias_stats],
            'counts': [count for _, count in categorias_stats]
        }

        # ========================================
        # 5. SPARKLINES - TENDÊNCIAS DOS ÚLTIMOS 12 MESES
        # ========================================
        sparklines = {}

        # Relatórios por mês (últimos 12)
        rel_sparkline = []
        for i in range(11, -1, -1):
            mes_inicio = datetime.utcnow() - timedelta(days=30*i)
            mes_fim = datetime.utcnow() - timedelta(days=30*(i-1)) if i > 0 else datetime.utcnow()
            count = Relatorio.query.filter(
                Relatorio.created_at >= mes_inicio,
                Relatorio.created_at < mes_fim
            ).count()
            rel_sparkline.append(count)
        sparklines['relatorios'] = rel_sparkline

        # Taxa de aprovação por mês (últimos 12)
        taxa_sparkline = []
        for i in range(11, -1, -1):
            mes_inicio = datetime.utcnow() - timedelta(days=30*i)
            mes_fim = datetime.utcnow() - timedelta(days=30*(i-1)) if i > 0 else datetime.utcnow()
            total = Prova.query.filter(
                Prova.created_at >= mes_inicio,
                Prova.created_at < mes_fim
            ).count()
            aprovadas = Prova.query.filter(
                Prova.created_at >= mes_inicio,
                Prova.created_at < mes_fim,
                Prova.status == 'APROVADA'
            ).count()
            taxa = round((aprovadas / total * 100) if total > 0 else 0, 1)
            taxa_sparkline.append(taxa)
        sparklines['aprovacao'] = taxa_sparkline

        # ========================================
        # 6. GRÁFICO MISTO - PROVAS E TAXA DE APROVAÇÃO POR MÊS
        # ========================================
        mixed_data = {
            'labels': [],
            'totalProvas': [],
            'taxaAprovacao': []
        }

        for i in range(5, -1, -1):
            mes_inicio = datetime.utcnow() - timedelta(days=30*i)
            mes_fim = datetime.utcnow() - timedelta(days=30*(i-1)) if i > 0 else datetime.utcnow()

            # Total de provas
            total = Prova.query.filter(
                Prova.created_at >= mes_inicio,
                Prova.created_at < mes_fim
            ).count()

            # Provas aprovadas
            aprovadas = Prova.query.filter(
                Prova.created_at >= mes_inicio,
                Prova.created_at < mes_fim,
                Prova.status == 'APROVADA'
            ).count()

            # Taxa
            taxa = round((aprovadas / total * 100) if total > 0 else 0, 1)

            # Label do mês
            mes_label = meses_map.get(mes_inicio.strftime('%m'), mes_inicio.strftime('%b'))

            mixed_data['labels'].append(mes_label)
            mixed_data['totalProvas'].append(total)
            mixed_data['taxaAprovacao'].append(taxa)

        # ========================================
        # 7. COLEÇÕES
        # ========================================
        colecoes_stats = db.session.query(
            Relatorio.colecao,
            db.func.count(Relatorio.id)
        ).filter(
            Relatorio.colecao.isnot(None),
            Relatorio.colecao != ''
        ).group_by(
            Relatorio.colecao
        ).order_by(
            db.func.count(Relatorio.id).desc()
        ).limit(8).all()

        colecoes_chart = {
            'labels': [col for col, _ in colecoes_stats],
            'counts': [count for _, count in colecoes_stats]
        }

        # ========================================
        # RETORNAR JSON
        # ========================================
        return jsonify({
            'success': True,
            'data': {
                'statusChart': status_chart,
                'suppliersChart': suppliers_chart,
                'timelineChart': timeline_chart,
                'categoryChart': category_chart,
                'sparklines': sparklines,
                'mixedChart': mixed_data,
                'colecoesChart': colecoes_chart
            }
        })

    except Exception as e:
        app.logger.error(f"Erro ao gerar dados dos gráficos: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/analytics/exportar')
@login_required
def analytics_exportar():
    """Exporta dados filtrados para Excel"""
    from datetime import datetime, timedelta
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Obter parâmetros de filtro
    filtro_status = request.args.get('status', '')
    filtro_categoria = request.args.get('categoria', '')
    filtro_colecao = request.args.get('colecao', '')
    filtro_fornecedor = request.args.get('fornecedor', '')
    filtro_referencia = request.args.get('referencia', '')
    filtro_data_inicio = request.args.get('data_inicio', '')
    filtro_data_fim = request.args.get('data_fim', '')

    # Query base
    query_provas = Prova.query.join(Referencia).join(Relatorio)

    # Aplicar filtros
    if filtro_status:
        query_provas = query_provas.filter(Prova.status == filtro_status)
    if filtro_categoria:
        query_provas = query_provas.filter(Referencia.tipo_categoria == filtro_categoria)
    if filtro_colecao:
        query_provas = query_provas.filter(Relatorio.colecao == filtro_colecao)
    if filtro_fornecedor:
        query_provas = query_provas.filter(Referencia.fornecedor == filtro_fornecedor)
    if filtro_referencia:
        # Busca parcial por número de referência (case-insensitive)
        query_provas = query_provas.filter(Referencia.numero_ref.ilike(f'%{filtro_referencia}%'))
    if filtro_data_inicio:
        try:
            data_inicio = datetime.strptime(filtro_data_inicio, '%Y-%m-%d')
            query_provas = query_provas.filter(Prova.created_at >= data_inicio)
        except:
            pass
    if filtro_data_fim:
        try:
            data_fim = datetime.strptime(filtro_data_fim, '%Y-%m-%d')
            data_fim = data_fim + timedelta(days=1)
            query_provas = query_provas.filter(Prova.created_at < data_fim)
        except:
            pass

    provas = query_provas.order_by(Relatorio.colecao, Referencia.tipo_categoria, Prova.numero_prova).all()

    # Criar workbook
    wb = Workbook()

    # ========================================
    # ABA 1: RESUMO
    # ========================================
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="e6007e", end_color="e6007e", fill_type="solid")
    subheader_fill = PatternFill(start_color="f8bbd9", end_color="f8bbd9", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws_resumo['A1'] = "RELATÓRIO DE ANALYTICS - PROVAS DE MODELAGEM"
    ws_resumo['A1'].font = Font(bold=True, size=14, color="e6007e")
    ws_resumo.merge_cells('A1:D1')

    ws_resumo['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_resumo['A2'].font = Font(italic=True, size=10)

    # Filtros aplicados
    ws_resumo['A4'] = "Filtros Aplicados:"
    ws_resumo['A4'].font = Font(bold=True)

    row = 5
    if filtro_status:
        ws_resumo[f'A{row}'] = f"Status: {filtro_status}"
        row += 1
    if filtro_categoria:
        ws_resumo[f'A{row}'] = f"Categoria: {filtro_categoria}"
        row += 1
    if filtro_colecao:
        ws_resumo[f'A{row}'] = f"Coleção: {filtro_colecao}"
        row += 1
    if filtro_fornecedor:
        ws_resumo[f'A{row}'] = f"Fornecedor: {filtro_fornecedor}"
        row += 1
    if filtro_data_inicio or filtro_data_fim:
        periodo = f"Período: {filtro_data_inicio or 'início'} até {filtro_data_fim or 'hoje'}"
        ws_resumo[f'A{row}'] = periodo
        row += 1

    if row == 5:
        ws_resumo[f'A{row}'] = "Nenhum filtro aplicado (dados completos)"
        row += 1

    # Estatísticas
    row += 1
    ws_resumo[f'A{row}'] = "Estatísticas"
    ws_resumo[f'A{row}'].font = Font(bold=True, size=12)

    row += 1
    total = len(provas)
    aprovadas = sum(1 for p in provas if p.status == 'APROVADA')
    reprovadas = sum(1 for p in provas if p.status == 'REPROVADA')
    em_andamento = sum(1 for p in provas if p.status == 'EM ANDAMENTO')
    comite = sum(1 for p in provas if p.status == 'COMITÊ')

    stats = [
        ("Total de Provas", total),
        ("Aprovadas", aprovadas),
        ("Reprovadas", reprovadas),
        ("Em Andamento", em_andamento),
        ("Comitê", comite),
        ("Taxa de Aprovação", f"{round(aprovadas/total*100, 1) if total > 0 else 0}%")
    ]

    for stat_name, stat_value in stats:
        ws_resumo[f'A{row}'] = stat_name
        ws_resumo[f'B{row}'] = stat_value
        row += 1

    # ========================================
    # ABA 2: DADOS DETALHADOS
    # ========================================
    ws_dados = wb.create_sheet("Dados Detalhados")

    headers = ["ID Relatório", "Coleção", "Descrição", "Referência", "Categoria",
               "Fornecedor", "Nº Prova", "Status", "Data Prova", "Data Recebimento",
               "Tamanhos", "Time Qualidade", "Time Estilo", "Time Modelagem"]

    for col_num, header in enumerate(headers, 1):
        cell = ws_dados.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_num, prova in enumerate(provas, 2):
        ref = prova.referencia
        rel = ref.relatorio

        data = [
            rel.id,
            rel.colecao or '',
            rel.descricao_geral or '',
            ref.numero_ref or '',
            ref.tipo_categoria,
            ref.fornecedor or '',
            prova.numero_prova,
            prova.status,
            prova.data_prova or '',
            prova.data_recebimento or '',
            prova.tamanhos_recebidos or '',
            prova.time_qualidade or '',
            prova.time_estilo or '',
            prova.time_modelagem or ''
        ]

        for col_num, value in enumerate(data, 1):
            cell = ws_dados.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border

            # Colorir por status
            if col_num == 8:  # Status
                if value == 'APROVADA':
                    cell.fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
                elif value == 'REPROVADA':
                    cell.fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
                elif value == 'COMITÊ':
                    cell.fill = PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid")

    # Ajustar larguras
    for ws in [ws_resumo, ws_dados]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Salvar
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"analytics_export_{timestamp}.xlsx"
    filepath = os.path.join(app.config['PDF_FOLDER'], filename)
    wb.save(filepath)

    return send_from_directory(app.config['PDF_FOLDER'], filename, as_attachment=True)


# ========================================
# LOGS DE AUDITORIA (Admin Only)
# ========================================

@app.route('/logs')
@login_required
def logs():
    """Página de visualização de logs de auditoria - Apenas para administradores"""
    # Verificar se usuário é admin
    if not current_user.is_admin and current_user.role != 'admin':
        flash('Acesso negado. Apenas administradores podem visualizar logs.', 'danger')
        return redirect(url_for('dashboard'))

    # Parâmetros de paginação e filtro
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    filtro_acao = request.args.get('acao', '')
    filtro_usuario = request.args.get('usuario', '')
    filtro_entidade = request.args.get('entidade', '')

    # Query base
    query = AuditLog.query

    # Aplicar filtros
    if filtro_acao:
        query = query.filter(AuditLog.acao == filtro_acao)
    if filtro_usuario:
        query = query.filter(AuditLog.usuario_nome.ilike(f'%{filtro_usuario}%'))
    if filtro_entidade:
        query = query.filter(AuditLog.entidade_tipo == filtro_entidade)

    # Ordenar por data (mais recente primeiro)
    query = query.order_by(desc(AuditLog.created_at))

    # Paginar resultados
    logs_paginados = query.paginate(page=page, per_page=per_page, error_out=False)

    # Estatísticas
    total_logs = AuditLog.query.count()
    acoes_unicos = db.session.query(AuditLog.acao, db.func.count(AuditLog.id)).group_by(AuditLog.acao).all()
    usuarios_ativos = db.session.query(AuditLog.usuario_nome, db.func.count(AuditLog.id)).group_by(AuditLog.usuario_nome).order_by(desc(db.func.count(AuditLog.id))).limit(10).all()

    return render_template('logs.html',
                         logs=logs_paginados.items,
                         pagination=logs_paginados,
                         total_logs=total_logs,
                         acoes_unicos=acoes_unicos,
                         usuarios_ativos=usuarios_ativos,
                         filtro_acao=filtro_acao,
                         filtro_usuario=filtro_usuario,
                         filtro_entidade=filtro_entidade)


# ========================================
# HELPER: Gerar Código Sequencial (F1)
# ========================================

def gerar_codigo_relatorio():
    """Gera código sequencial REL-YYYY-NNN"""
    from datetime import datetime
    ano = datetime.utcnow().year
    ultimo = Relatorio.query.filter(
        Relatorio.codigo.like(f'REL-{ano}-%')
    ).order_by(desc(Relatorio.id)).first()

    if ultimo and ultimo.codigo:
        try:
            numero = int(ultimo.codigo.split('-')[-1]) + 1
        except (ValueError, IndexError):
            numero = 1
    else:
        numero = 1

    return f'REL-{ano}-{numero:03d}'


# ========================================
# F1: DUPLICAR RELATÓRIO
# ========================================

@app.route('/relatorio/<int:id>/duplicar', methods=['POST'])
@login_required
def duplicar_relatorio(id):
    """Duplica um relatório com suas referências (sem provas)"""
    relatorio_original = Relatorio.query.get_or_404(id)

    try:
        novo_codigo = gerar_codigo_relatorio()

        novo_rel = Relatorio(
            codigo=novo_codigo,
            descricao_geral=f"[CÓPIA] {relatorio_original.descricao_geral}",
            colecao=relatorio_original.colecao,
            temporada=relatorio_original.temporada,
            ano=relatorio_original.ano,
            linha=relatorio_original.linha,
            status_geral='Em Andamento',
            data_limite=relatorio_original.data_limite,
            created_by=current_user.id if current_user.is_authenticated else None
        )
        db.session.add(novo_rel)
        db.session.flush()

        # Copiar referências (sem provas)
        for ref in relatorio_original.referencias:
            nova_ref = Referencia(
                relatorio_id=novo_rel.id,
                tipo_categoria=ref.tipo_categoria,
                numero_ref=ref.numero_ref,
                codigo_referencia=ref.codigo_referencia,
                origem=ref.origem,
                fornecedor=ref.fornecedor,
                fornecedor_contato=ref.fornecedor_contato,
                fornecedor_id=ref.fornecedor_id,
                materia_prima=ref.materia_prima,
                composicao=ref.composicao,
                gramatura=ref.gramatura,
                aviamentos=ref.aviamentos,
                observacoes=ref.observacoes
            )
            db.session.add(nova_ref)

        db.session.commit()

        registrar_log(
            acao='duplicar',
            entidade_tipo='relatorio',
            entidade_id=novo_rel.id,
            entidade_descricao=f'Duplicado de #{id}: {relatorio_original.descricao_geral}',
            detalhes=f'Código: {novo_codigo}'
        )

        flash(f"Relatório duplicado com sucesso! Novo código: {novo_codigo}", "success")
        return redirect(url_for('detalhes_relatorio', id=novo_rel.id))

    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao duplicar relatório: {e}", "error")
        return redirect(url_for('detalhes_relatorio', id=id))


# ========================================
# DUPLICAR PROVA
# ========================================

@app.route('/prova/<int:prova_id>/duplicar', methods=['POST'])
@login_required
def duplicar_prova(prova_id):
    """Duplica uma prova (sem fotos e sem lacre), gerando novo número e status 'Em Andamento'."""
    prova_original = Prova.query.get_or_404(prova_id)
    referencia = Referencia.query.get_or_404(prova_original.referencia_id)

    try:
        # Próximo número de prova para a mesma referência
        ultimo_numero = (
            db.session.query(db.func.max(Prova.numero_prova))
            .filter(Prova.referencia_id == referencia.id)
            .scalar()
        ) or 0
        novo_numero = ultimo_numero + 1

        nova_prova = Prova(
            referencia_id=referencia.id,
            numero_prova=novo_numero,
            status='Em Andamento',
            data_recebimento=prova_original.data_recebimento,
            tamanhos_recebidos=prova_original.tamanhos_recebidos,
            info_medidas=prova_original.info_medidas,
            data_prova=None,
            time_qualidade=prova_original.time_qualidade,
            checklist_qualidade=prova_original.checklist_qualidade,
            comentarios_qualidade=prova_original.comentarios_qualidade,
            obs_qualidade=prova_original.obs_qualidade,
            time_estilo=prova_original.time_estilo,
            checklist_estilo=prova_original.checklist_estilo,
            comentarios_estilo=prova_original.comentarios_estilo,
            obs_estilo=prova_original.obs_estilo,
            time_modelagem=prova_original.time_modelagem,
            checklist_modelagem=prova_original.checklist_modelagem,
            comentarios_modelagem=prova_original.comentarios_modelagem,
            obs_modelagem=prova_original.obs_modelagem,
            info_adicionais=prova_original.info_adicionais,
        )
        db.session.add(nova_prova)
        db.session.commit()

        registrar_log(
            acao='duplicar',
            entidade_tipo='prova',
            entidade_id=nova_prova.id,
            entidade_descricao=f'Duplicada da prova #{prova_original.id} (nº {prova_original.numero_prova})',
            detalhes=f'Nova prova nº {novo_numero} na referência {referencia.id}'
        )

        flash(f'Prova duplicada com sucesso! Nova prova nº {novo_numero}.', 'success')
        return redirect(url_for('editar_relatorio', id=referencia.relatorio_id))

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao duplicar prova: {e}', 'danger')
        return redirect(url_for('detalhes_relatorio', id=referencia.relatorio_id))


# ========================================
# F5: COMPARAÇÃO DE PROVAS
# ========================================

@app.route('/referencia/<int:id>/comparar')
@login_required
def comparar_provas(id):
    """Compara duas provas lado a lado"""
    referencia = Referencia.query.get_or_404(id)
    prova1_id = request.args.get('prova1', type=int)
    prova2_id = request.args.get('prova2', type=int)

    if not prova1_id or not prova2_id:
        flash("Selecione duas provas para comparar.", "warning")
        return redirect(url_for('detalhes_relatorio', id=referencia.relatorio_id))

    prova1 = Prova.query.get_or_404(prova1_id)
    prova2 = Prova.query.get_or_404(prova2_id)

    # Prepare prova dicts with photos
    def prova_to_dict(prova):
        prova_dict = {c.name: getattr(prova, c.name) for c in prova.__table__.columns}
        prova_dict['fotos'] = {}
        for foto in prova.fotos:
            contexto = foto.contexto
            if contexto not in prova_dict['fotos']:
                prova_dict['fotos'][contexto] = []
            prova_dict['fotos'][contexto].append({c.name: getattr(foto, c.name) for c in foto.__table__.columns})

        # Checklist respostas
        respostas = ChecklistResposta.query.filter_by(prova_id=prova.id).all()
        prova_dict['checklist_respostas'] = respostas
        total = len(respostas)
        conformes = sum(1 for r in respostas if r.conforme)
        prova_dict['checklist_conformidade'] = round(conformes / total * 100, 1) if total > 0 else 0

        return prova_dict

    prova1_dict = prova_to_dict(prova1)
    prova2_dict = prova_to_dict(prova2)

    # Find differences
    campos_comparar = [
        'status', 'data_recebimento', 'tamanhos_recebidos', 'data_prova',
        'time_qualidade', 'comentarios_qualidade', 'obs_qualidade',
        'time_estilo', 'comentarios_estilo', 'obs_estilo',
        'time_modelagem', 'comentarios_modelagem', 'obs_modelagem',
        'data_lacre', 'numero_lacre', 'info_adicionais'
    ]
    diferencas = {}
    for campo in campos_comparar:
        v1 = prova1_dict.get(campo) or ''
        v2 = prova2_dict.get(campo) or ''
        if v1 != v2:
            diferencas[campo] = True

    return render_template('comparar_provas.html',
                           referencia=referencia,
                           prova1=prova1_dict,
                           prova2=prova2_dict,
                           diferencas=diferencas)


# ========================================
# F7: ANALYTICS PDF (Relatório Gerencial)
# ========================================

@app.route('/analytics/pdf')
@login_required
def analytics_pdf():
    """Gera PDF gerencial com métricas de desempenho"""
    from datetime import datetime, timedelta

    total_relatorios = Relatorio.query.count()
    total_referencias = Referencia.query.count()
    total_provas = Prova.query.count()

    provas_aprovadas = Prova.query.filter_by(status='APROVADA').count()
    provas_reprovadas = Prova.query.filter_by(status='REPROVADA').count()
    provas_em_andamento = Prova.query.filter_by(status='EM ANDAMENTO').count()
    provas_comite = Prova.query.filter_by(status='COMITÊ').count()

    taxa_aprovacao = round((provas_aprovadas / total_provas * 100) if total_provas > 0 else 0, 1)
    provas_retrabalho = Prova.query.filter(Prova.numero_prova > 1).count()
    taxa_retrabalho = round((provas_retrabalho / total_provas * 100) if total_provas > 0 else 0, 1)

    # Por categoria
    categorias_stats = db.session.query(
        Referencia.tipo_categoria,
        db.func.count(Referencia.id)
    ).group_by(Referencia.tipo_categoria).all()

    # Top fornecedores
    fornecedores_stats = db.session.query(
        Referencia.fornecedor,
        db.func.count(Referencia.id)
    ).filter(Referencia.fornecedor.isnot(None), Referencia.fornecedor != '').group_by(
        Referencia.fornecedor
    ).order_by(db.func.count(Referencia.id).desc()).limit(10).all()

    # Relatórios com prazo vencido
    from datetime import date
    hoje = date.today().isoformat()
    relatorios_vencidos = Relatorio.query.filter(
        Relatorio.data_limite.isnot(None),
        Relatorio.data_limite != '',
        Relatorio.data_limite < hoje,
        Relatorio.status_geral != 'Aprovado'
    ).count()

    html_string = render_template('analytics_pdf.html',
                                   total_relatorios=total_relatorios,
                                   total_referencias=total_referencias,
                                   total_provas=total_provas,
                                   provas_aprovadas=provas_aprovadas,
                                   provas_reprovadas=provas_reprovadas,
                                   provas_em_andamento=provas_em_andamento,
                                   provas_comite=provas_comite,
                                   taxa_aprovacao=taxa_aprovacao,
                                   taxa_retrabalho=taxa_retrabalho,
                                   categorias_stats=dict(categorias_stats),
                                   fornecedores_stats=fornecedores_stats,
                                   relatorios_vencidos=relatorios_vencidos,
                                   now=datetime.now)

    pdf = HTML(string=html_string).write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=analytics_relatorio_{datetime.now().strftime("%Y%m%d")}.pdf'
    return response


# ========================================
# F8: VERSIONAMENTO DE ARQUIVOS
# ========================================

@app.route('/arquivo/<int:entidade_id>/versoes')
@login_required
def listar_versoes(entidade_id):
    """Lista versões de arquivo de uma entidade"""
    entidade_tipo = request.args.get('tipo', 'relatorio')
    campo = request.args.get('campo', '')

    versoes = ArquivoVersao.query.filter_by(
        entidade_tipo=entidade_tipo,
        entidade_id=entidade_id,
        campo=campo
    ).order_by(desc(ArquivoVersao.versao)).all()

    return jsonify([{
        'id': v.id,
        'versao': v.versao,
        'file_path': v.file_path,
        'created_at': v.created_at.strftime('%d/%m/%Y %H:%M') if v.created_at else '',
    } for v in versoes])


@app.route('/arquivo/versao/<int:id>/download')
@login_required
def download_versao(id):
    """Download de uma versão específica de arquivo"""
    versao = ArquivoVersao.query.get_or_404(id)
    return send_from_directory(app.config['UPLOAD_FOLDER'], versao.file_path)


def salvar_versao_arquivo(entidade_tipo, entidade_id, campo, file_path_antigo):
    """Salva uma versão anterior do arquivo antes de substituir"""
    if not file_path_antigo:
        return

    try:
        # Contar versões existentes
        ultima_versao = ArquivoVersao.query.filter_by(
            entidade_tipo=entidade_tipo,
            entidade_id=entidade_id,
            campo=campo
        ).order_by(desc(ArquivoVersao.versao)).first()

        nova_versao = (ultima_versao.versao + 1) if ultima_versao else 1

        versao = ArquivoVersao(
            entidade_tipo=entidade_tipo,
            entidade_id=entidade_id,
            campo=campo,
            file_path=file_path_antigo,
            versao=nova_versao,
            uploaded_by=current_user.id if current_user.is_authenticated else None
        )
        db.session.add(versao)
    except Exception as e:
        app.logger.error(f'Erro ao salvar versão de arquivo: {e}')


# ========================================
# F9: DASHBOARD PERSONALIZADO
# ========================================

@app.route('/dashboard/preferencias', methods=['POST'])
@login_required
def salvar_preferencias_dashboard():
    """Salva configuração de widgets do dashboard (AJAX)"""
    try:
        config = request.get_json()
        if not config:
            return jsonify({'success': False, 'error': 'Dados inválidos'}), 400

        pref = PreferenciaUsuario.query.filter_by(usuario_id=current_user.id).first()
        if not pref:
            pref = PreferenciaUsuario(usuario_id=current_user.id)
            db.session.add(pref)

        pref.set_config(config)
        db.session.commit()

        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ========================================
# Comandos CLI para manutenção
# ========================================

@app.cli.command('reset-all-passwords')
def reset_all_passwords():
    """Reseta as senhas de todos os usuários para uma senha padrão temporária."""
    from models import User
    from werkzeug.security import generate_password_hash

    HASH_METHOD = 'pbkdf2:sha256'
    senha_padrao = 'mudar123'

    users = User.query.all()
    print(f"Resetando senhas de {len(users)} usuários...")

    for user in users:
        user.password_hash = generate_password_hash(senha_padrao, method=HASH_METHOD)
        user.senha_temporaria = True
        print(f"  - {user.username}: senha resetada")

    db.session.commit()
    print(f"\nTodas as senhas foram resetadas para: {senha_padrao}")
    print("Os usuários deverão trocar a senha no primeiro acesso.")


@app.cli.command('create-admin')
def create_admin():
    """Cria um usuário administrador padrão."""
    from models import User
    from werkzeug.security import generate_password_hash

    HASH_METHOD = 'pbkdf2:sha256'
    username = 'admin'
    senha = 'admin123'

    # Verificar se já existe
    if User.query.filter_by(username=username).first():
        print(f"Usuário '{username}' já existe.")
        return

    admin = User(
        username=username,
        email='admin@sistema.local',
        nome_completo='Administrador',
        password_hash=generate_password_hash(senha, method=HASH_METHOD),
        role='admin',
        is_admin=True,
        is_active=True,
        senha_temporaria=True
    )

    db.session.add(admin)
    db.session.commit()

    print(f"Usuário administrador criado!")
    print(f"  Username: {username}")
    print(f"  Senha: {senha}")
    print("  ** Troque a senha após o primeiro login! **")


if __name__ == '__main__':
    # Este bloco é usado apenas para desenvolvimento local
    # Em produção, use Gunicorn: gunicorn -c gunicorn_config.py wsgi:app
    app.run(
        debug=app.config.get('DEBUG', False),
        host=os.getenv('HOST', '127.0.0.1'),
        port=int(os.getenv('PORT', 5000))
    )
