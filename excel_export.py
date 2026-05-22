from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import current_app
import os
import tempfile
from datetime import datetime

def export_relatorios_to_excel(relatorios_data):
    """
    Exporta relatórios para arquivo Excel
    
    Args:
        relatorios_data: Lista de dicionários com dados dos relatórios
    
    Returns:
        str: Caminho do arquivo Excel gerado
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatórios"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="e6007e", end_color="e6007e", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Cabeçalhos expandidos
    headers = [
        "ID", "Código", "Coleção", "Descrição", "Temporada", "Ano",
        "Status Geral", "Referências", "Data Criação", "Última Atualização"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Dados
    for row_num, relatorio in enumerate(relatorios_data, 2):
        ws.cell(row=row_num, column=1, value=relatorio.get('id', ''))
        ws.cell(row=row_num, column=2, value=relatorio.get('codigo', ''))
        ws.cell(row=row_num, column=3, value=relatorio.get('colecao', ''))
        ws.cell(row=row_num, column=4, value=relatorio.get('descricao_geral', ''))
        ws.cell(row=row_num, column=5, value=relatorio.get('temporada', ''))
        ws.cell(row=row_num, column=6, value=relatorio.get('ano', ''))
        ws.cell(row=row_num, column=7, value=relatorio.get('status_geral', ''))
        ws.cell(row=row_num, column=8, value=relatorio.get('num_referencias', 0))
        ws.cell(row=row_num, column=9, value=relatorio.get('data_criacao', ''))
        ws.cell(row=row_num, column=10, value=relatorio.get('data_atualizacao', ''))
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar arquivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"relatorios_export_{timestamp}.xlsx"
    filepath = os.path.join(current_app.config['PDF_FOLDER'], filename)
    wb.save(filepath)
    
    return filename

def export_detalhes_to_excel(relatorio, referencias):
    """
    Exporta detalhes completos de um relatório para Excel
    
    Args:
        relatorio: Dados do relatório
        referencias: Lista de referências com provas
    
    Returns:
        str: Caminho do arquivo Excel gerado
    """
    wb = Workbook()
    
    # Aba 1: Informações Gerais
    ws_geral = wb.active
    ws_geral.title = "Informações Gerais"
    
    ws_geral['A1'] = "Campo"
    ws_geral['B1'] = "Valor"
    ws_geral['A1'].font = Font(bold=True)
    ws_geral['B1'].font = Font(bold=True)
    
    ws_geral['A2'] = "ID"
    ws_geral['B2'] = relatorio['id']
    ws_geral['A3'] = "Coleção"
    ws_geral['B3'] = relatorio.get('colecao', '')
    ws_geral['A4'] = "Descrição"
    ws_geral['B4'] = relatorio.get('descricao_geral', '')
    
    # Aba 2: Referências
    ws_refs = wb.create_sheet("Referências")
    headers_refs = [
        "Referência", "Tipo", "Origem", "Fornecedor", "País Fornecedor", "Contato Fornecedor",
        "Matéria Prima", "Composição", "Gramatura", "Aviamentos", "Observações"
    ]
    for col_num, header in enumerate(headers_refs, 1):
        cell = ws_refs.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    row_num = 2
    for ref in referencias:
        ws_refs.cell(row=row_num, column=1, value=ref.get('numero_ref', ''))
        ws_refs.cell(row=row_num, column=2, value=ref.get('tipo', ''))
        ws_refs.cell(row=row_num, column=3, value=ref.get('origem', ''))
        ws_refs.cell(row=row_num, column=4, value=ref.get('fornecedor', ''))
        ws_refs.cell(row=row_num, column=5, value=ref.get('fornecedor_pais', ''))
        ws_refs.cell(row=row_num, column=6, value=ref.get('fornecedor_contato', ''))
        ws_refs.cell(row=row_num, column=7, value=ref.get('materia_prima', ''))
        ws_refs.cell(row=row_num, column=8, value=ref.get('composicao', ''))
        ws_refs.cell(row=row_num, column=9, value=ref.get('gramatura', ''))
        ws_refs.cell(row=row_num, column=10, value=ref.get('aviamentos', ''))
        ws_refs.cell(row=row_num, column=11, value=ref.get('observacoes', ''))
        row_num += 1

    # Aba 3: Provas
    ws_provas = wb.create_sheet("Provas")

    headers_provas = [
        "Referência", "Tipo", "Nº Prova", "Status", "Data Recebimento", "Data Prova",
        "Tamanhos", "Info Medidas", "Time Qualidade", "Time Estilo", "Time Modelagem",
        "Data Lacre", "Nº Lacre", "Informações Adicionais"
    ]
    for col_num, header in enumerate(headers_provas, 1):
        cell = ws_provas.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    row_num = 2
    for ref in referencias:
        for prova in ref.get('provas', []):
            ws_provas.cell(row=row_num, column=1, value=ref.get('numero_ref', ''))
            ws_provas.cell(row=row_num, column=2, value=ref.get('tipo', ''))
            ws_provas.cell(row=row_num, column=3, value=prova.get('numero_prova', ''))
            ws_provas.cell(row=row_num, column=4, value=prova.get('status', ''))
            ws_provas.cell(row=row_num, column=5, value=prova.get('data_recebimento', ''))
            ws_provas.cell(row=row_num, column=6, value=prova.get('data_prova', ''))
            ws_provas.cell(row=row_num, column=7, value=prova.get('tamanhos_recebidos', ''))
            ws_provas.cell(row=row_num, column=8, value=prova.get('info_medidas', ''))
            ws_provas.cell(row=row_num, column=9, value=prova.get('time_qualidade', ''))
            ws_provas.cell(row=row_num, column=10, value=prova.get('time_estilo', ''))
            ws_provas.cell(row=row_num, column=11, value=prova.get('time_modelagem', ''))
            ws_provas.cell(row=row_num, column=12, value=prova.get('data_lacre', ''))
            ws_provas.cell(row=row_num, column=13, value=prova.get('numero_lacre', ''))
            ws_provas.cell(row=row_num, column=14, value=prova.get('info_adicionais', ''))
            row_num += 1
    
    # Ajustar larguras
    for ws in [ws_geral, ws_refs, ws_provas]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"relatorio_{relatorio['id']}_detalhes_{timestamp}.xlsx"
    filepath = os.path.join(current_app.config['PDF_FOLDER'], filename)
    wb.save(filepath)

    return filename


def export_editavel(relatorios):
    """
    Exporta relatórios no MESMO formato aceito por /importar/excel,
    incluindo colunas de ID (ocultas) para suportar round-trip
    (export → editar → reimportar atualizando registros existentes).

    Args:
        relatorios: lista de objetos Relatorio (SQLAlchemy) com referências e provas

    Returns:
        str: caminho absoluto do arquivo .xlsx gerado (temporário)
    """
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="e6007e", end_color="e6007e", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    id_fill = PatternFill(start_color="f0f0f0", end_color="f0f0f0", fill_type="solid")

    # ABA 1 — Informações Gerais
    ws1 = wb.active
    ws1.title = "Informações Gerais"
    headers1 = ['ID Relatorio', 'Descrição', 'Linha', 'Coleção', 'Temporada', 'Ano', 'Status Geral']
    for col_idx, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill if col_idx > 1 else id_fill
        c.alignment = header_alignment
        c.border = thin_border

    row = 2
    for rel in relatorios:
        ws1.cell(row=row, column=1, value=rel.id).fill = id_fill
        ws1.cell(row=row, column=2, value=rel.descricao_geral)
        ws1.cell(row=row, column=3, value=rel.linha)
        ws1.cell(row=row, column=4, value=rel.colecao)
        ws1.cell(row=row, column=5, value=rel.temporada)
        ws1.cell(row=row, column=6, value=rel.ano)
        ws1.cell(row=row, column=7, value=rel.status_geral)
        for col_idx in range(1, 8):
            ws1.cell(row=row, column=col_idx).alignment = cell_alignment
        row += 1

    col_widths1 = [12, 30, 15, 20, 15, 10, 18]
    for col_idx, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(col_idx)].width = w

    # ABA 2 — Dados Detalhados
    ws2 = wb.create_sheet("Dados Detalhados")
    headers2 = [
        'ID Relatorio', 'ID Referencia', 'ID Prova',
        'Coleção', 'Descrição', 'Referência', 'Categoria',
        'Fornecedor', 'Nº Prova', 'Status', 'Data Prova',
        'Data Recebimento', 'Tamanhos', 'Time Qualidade',
        'Time Estilo', 'Time Modelagem',
    ]
    for col_idx, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill if col_idx > 3 else id_fill
        c.alignment = header_alignment
        c.border = thin_border

    row = 2
    for rel in relatorios:
        for ref in rel.referencias:
            # Se a referência não tem provas, ainda inserir 1 linha com IDs
            provas = list(ref.provas) if ref.provas else [None]
            for prova in provas:
                ws2.cell(row=row, column=1, value=rel.id).fill = id_fill
                ws2.cell(row=row, column=2, value=ref.id).fill = id_fill
                if prova is not None:
                    ws2.cell(row=row, column=3, value=prova.id).fill = id_fill
                else:
                    ws2.cell(row=row, column=3, value='').fill = id_fill
                ws2.cell(row=row, column=4, value=rel.colecao)
                ws2.cell(row=row, column=5, value=rel.descricao_geral)
                ws2.cell(row=row, column=6, value=ref.numero_ref)
                ws2.cell(row=row, column=7, value=ref.tipo_categoria)
                fornecedor_nome = None
                if hasattr(ref, 'fornecedor_obj') and ref.fornecedor_obj:
                    fornecedor_nome = ref.fornecedor_obj.nome
                else:
                    fornecedor_nome = ref.fornecedor
                ws2.cell(row=row, column=8, value=fornecedor_nome)
                if prova is not None:
                    ws2.cell(row=row, column=9, value=prova.numero_prova)
                    ws2.cell(row=row, column=10, value=prova.status)
                    ws2.cell(row=row, column=11, value=prova.data_prova)
                    ws2.cell(row=row, column=12, value=prova.data_recebimento)
                    ws2.cell(row=row, column=13, value=prova.tamanhos_recebidos)
                    ws2.cell(row=row, column=14, value=prova.time_qualidade)
                    ws2.cell(row=row, column=15, value=prova.time_estilo)
                    ws2.cell(row=row, column=16, value=prova.time_modelagem)
                for col_idx in range(1, 17):
                    ws2.cell(row=row, column=col_idx).alignment = cell_alignment
                row += 1

    col_widths2 = [12, 12, 10, 20, 30, 15, 12, 20, 10, 16, 14, 16, 15, 18, 18, 18]
    for col_idx, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    # Salvar em arquivo temporário
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp.close()
    wb.save(tmp.name)
    return tmp.name
